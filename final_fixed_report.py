#!/usr/bin/env python3
"""
ОКОНЧАТЕЛЬНО ИСПРАВЛЕННАЯ версия отчета.

ФИЛЬТРАЦИЯ И ПОДСЧЁТ:

Форма 2304918 (Старички - возврат студентов):
  * Статус PE: Start, Future, PE 5, Китайский
  * Даты выхода (поля 26, 31, 56):
    - Если ВСЕ пустые → ВКЛЮЧАЕМ (не вышел на обучение)
    - Если хотя бы одна заполнена → ВСЕ заполненные даты из августа-сентября 2025
  * Учится: поле 64 (чекбокс)
  * Китайский: учитывается в личном зачёте, НЕ учитывается в статистике филиалов
  * Включая закрытые формы
  
Форма 792300 (БПЗ - конверсия после пробного):
  * БАЗОВОЕ КОЛИЧЕСТВО (всего):
    - Статус PE: Start, Future, PE 5, Китайский (проверяется ПЕРВЫМ)
    - Поле 220 (дата БПЗ) = август-сентябрь 2025 (обязательна!)
    - Поле 183 = "Да" (подтверждение прихода на БПЗ)
    - Включая закрытые формы
    - Группировка по клиентам (ФИО + филиал), логика "хотя бы один"
  * УЧИТСЯ (для процента):
    - Из клиентов базы: хотя бы одна форма с 183="Да" + поле 181="Сентябрь"
  * Китайский: учитывается и в личном, и в филиальном зачёте
"""

import asyncio
import sys
import os
import json
from pathlib import Path
from typing import Dict, List, Any, Optional, Set
from datetime import datetime
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv

# Загружаем переменные окружения из .env файла
load_dotenv()

# Добавляем app в путь для импорта
sys.path.append(str(Path(__file__).parent / "app"))

from pyrus_client import PyrusClient


class TeacherStats:
    """Статистика по преподавателю."""
    
    def __init__(self, name: str):
        self.name = name
        # Форма 2304918 (возврат студентов)
        self.form_2304918_total = 0
        self.form_2304918_studying = 0
        self.form_2304918_data = []  # Исходные данные для проверки
        
        # Форма 792300 (конверсия после БПЗ)
        self.form_792300_total = 0
        self.form_792300_studying = 0
        self.form_792300_data = []  # Исходные данные для проверки
    
    @property
    def return_percentage(self) -> float:
        """Процент возврата студентов (форма 2304918)."""
        if self.form_2304918_total == 0:
            return 0.0
        return (self.form_2304918_studying / self.form_2304918_total) * 100
    
    @property
    def conversion_percentage(self) -> float:
        """Процент конверсии после БПЗ → студент (форма 792300)."""
        if self.form_792300_total == 0:
            return 0.0
        return (self.form_792300_studying / self.form_792300_total) * 100
    
    @property
    def total_percentage(self) -> float:
        """Суммарный процент двух показателей."""
        return self.return_percentage + self.conversion_percentage


class BranchStats:
    """Статистика по филиалу."""
    
    def __init__(self, name: str):
        self.name = name
        # Форма 2304918 (старички)
        self.form_2304918_total = 0
        self.form_2304918_studying = 0
        
        # Форма 792300 (новый клиент)
        self.form_792300_total = 0
        self.form_792300_studying = 0
    
    @property
    def return_percentage(self) -> float:
        """Процент возврата студентов (старички)."""
        if self.form_2304918_total == 0:
            return 0.0
        return (self.form_2304918_studying / self.form_2304918_total) * 100
    
    @property
    def conversion_percentage(self) -> float:
        """Процент конверсии после БПЗ → студент (новый клиент)."""
        if self.form_792300_total == 0:
            return 0.0
        return (self.form_792300_studying / self.form_792300_total) * 100
    
    @property
    def total_percentage(self) -> float:
        """Суммарный процент двух показателей."""
        return self.return_percentage + self.conversion_percentage


class FinalFixedPyrusDataAnalyzer:
    """ОКОНЧАТЕЛЬНО исправленный анализатор данных из Pyrus."""
    
    def __init__(self):
        self.client = PyrusClient()
        self.teachers_stats: Dict[str, TeacherStats] = {}
        self.branches_stats: Dict[str, BranchStats] = {}
        
        # Загружаем списки исключений преподавателей
        self.excluded_teachers = self._load_exclusions()
        
        # Отладочные счетчики
        self.debug_target = "Анастасия Алексеевна Нечунаева"
        self.debug_counters = {
            "2304918_found": 0,
            "2304918_valid_pe": 0,
            "2304918_valid_dates": 0,
            "2304918_excluded": 0,
            "2304918_processed": 0,
            "792300_found": 0,
            "792300_valid_pe": 0,
            "792300_valid_dates": 0,
            "792300_excluded": 0,
            "792300_processed": 0
        }
    
    def _get_field_value(self, field_list: List[Dict[str, Any]], field_id: int) -> Optional[Any]:
        """Ищет значение поля по id, рекурсивно обходя вложенные секции."""
        for f in field_list or []:
            if f.get("id") == field_id:
                return f.get("value")
            val = f.get("value")
            if isinstance(val, dict) and isinstance(val.get("fields"), list):
                nested = self._get_field_value(val.get("fields") or [], field_id)
                if nested is not None:
                    return nested
        return None
    
    def _extract_teacher_name(self, task_fields: List[Dict[str, Any]], field_id: int) -> str:
        """Извлекает ФИО преподавателя из поля справочника."""
        value = self._get_field_value(task_fields, field_id)
        
        if isinstance(value, dict):
            # Поддержка person-объекта: first_name/last_name
            first_name = value.get("first_name", "")
            last_name = value.get("last_name", "")
            if isinstance(first_name, str) or isinstance(last_name, str):
                full_name = f"{(first_name or '').strip()} {(last_name or '').strip()}".strip()
                if full_name:
                    return full_name
            
            # Для справочника сотрудников обычно есть поле text или name
            for key in ("text", "name", "value"):
                name_val = value.get(key)
                if isinstance(name_val, str) and name_val.strip():
                    return name_val.strip()
        
        if isinstance(value, str):
            return value.strip()
        
        return "Неизвестный преподаватель"
    
    def _load_exclusions(self) -> Dict[str, Set[str]]:
        """Загружает списки исключений преподавателей из JSON файла."""
        exclusions_file = Path(__file__).parent / "teacher_exclusions.json"
        
        try:
            with open(exclusions_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return {
                    'oldies': set(data.get('oldies', [])),
                    'trial': set(data.get('trial', []))
                }
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"⚠️ Не удалось загрузить исключения из {exclusions_file}: {e}")
            print("Продолжаем без фильтрации преподавателей.")
            return {'oldies': set(), 'trial': set()}
    
    def _is_teacher_excluded(self, teacher_name: str, form_type: str) -> bool:
        """Проверяет, исключен ли преподаватель из указанной формы."""
        excluded_set = self.excluded_teachers.get(form_type, set())
        
        # Проверяем точное совпадение
        if teacher_name in excluded_set:
            return True
        
        # Проверяем частичное совпадение (фамилия входит в имя преподавателя)
        for excluded_name in excluded_set:
            if excluded_name.lower() in teacher_name.lower():
                return True
        
        return False
    
    def _normalize_branch_name(self, branch_name: str) -> str:
        """Нормализует название филиала для объединения данных."""
        branch_name = branch_name.lower().strip()
        
        # Только филиал Коммунистический 22 считается как Копейск
        if "коммунистический" in branch_name and "22" in branch_name:
            return "Копейск"
        # Славы 30 больше НЕ объединяется с Копейском
        
        # Возвращаем оригинальное название с заглавной буквы
        return branch_name.title()
    
    def _is_branch_excluded_from_competition(self, branch_name: str) -> bool:
        """Проверяет, исключен ли филиал из соревнования между филиалами."""
        branch_name = branch_name.lower().strip()
        
        # Исключаем из соревнования филиалов (но НЕ из статистики преподавателей!)
        if "макеева" in branch_name and "15" in branch_name:
            return True
        if "коммуны" in branch_name and "106/1" in branch_name:
            return True
        if "славы" in branch_name and "30" in branch_name:
            return True
        if "online" in branch_name or branch_name == "online":
            return True
        
        return False
    
    def _extract_branch_name(self, task_fields: List[Dict[str, Any]], field_id: int) -> str:
        """Извлекает название филиала из поля справочника."""
        value = self._get_field_value(task_fields, field_id)
        
        if isinstance(value, dict):
            # Проверяем массив values - основной способ для справочника филиалов
            values = value.get("values")
            if isinstance(values, list) and len(values) > 0:
                branch_name = values[0]  # Берем первое значение
                if isinstance(branch_name, str) and branch_name.strip():
                    return self._normalize_branch_name(branch_name.strip())
            
            # Проверяем rows если values не найден  
            rows = value.get("rows")
            if isinstance(rows, list) and len(rows) > 0 and isinstance(rows[0], list) and len(rows[0]) > 0:
                branch_name = rows[0][0]  # Берем первую ячейку первой строки
                if isinstance(branch_name, str) and branch_name.strip():
                    return self._normalize_branch_name(branch_name.strip())
            
            # Для обычных справочников обычно есть поле text или name
            for key in ("text", "name", "value"):
                branch_val = value.get(key)
                if isinstance(branch_val, str) and branch_val.strip():
                    return self._normalize_branch_name(branch_val.strip())
        
        if isinstance(value, str):
            return self._normalize_branch_name(value.strip())
        
        return "Неизвестный филиал"
    
    def _is_valid_pe_status(self, task_fields: List[Dict[str, Any]], field_id: int) -> bool:
        """Проверяет, соответствует ли статус PE одному из допустимых: PE Start, PE Future, PE 5, Китайский."""
        value = self._get_field_value(task_fields, field_id)
        
        # Допустимые статусы PE
        valid_statuses = {"PE Start", "PE Future", "PE 5", "Китайский"}
        
        if isinstance(value, dict):
            # Проверяем choice_names для справочника выбора
            choice_names = value.get("choice_names")
            if isinstance(choice_names, list) and len(choice_names) > 0:
                status = choice_names[0]
                if isinstance(status, str) and status.strip() in valid_statuses:
                    return True
            
            # Проверяем массив values для справочника
            values = value.get("values")
            if isinstance(values, list) and len(values) > 0:
                status = values[0]
                if isinstance(status, str) and status.strip() in valid_statuses:
                    return True
            
            # Проверяем rows если values не найден  
            rows = value.get("rows")
            if isinstance(rows, list) and len(rows) > 0 and isinstance(rows[0], list) and len(rows[0]) > 0:
                status = rows[0][0]
                if isinstance(status, str) and status.strip() in valid_statuses:
                    return True
            
            # Для обычных справочников проверяем text, name, value
            for key in ("text", "name", "value"):
                status_val = value.get(key)
                if isinstance(status_val, str) and status_val.strip() in valid_statuses:
                    return True
        
        if isinstance(value, str):
            return value.strip() in valid_statuses
        
        return False
    
    def _is_studying(self, task_fields: List[Dict[str, Any]], field_id: int) -> bool:
        """Проверяет, отмечена ли галочка 'учится' в указанном поле."""
        value = self._get_field_value(task_fields, field_id)
        
        if value is None:
            return False
        
        # Булево значение
        if isinstance(value, bool):
            return value
        
        # Строковое значение (checked/unchecked)
        if isinstance(value, str):
            return value.lower() in ("да", "yes", "true", "1", "checked")
        
        # Объект с чекбоксом
        if isinstance(value, dict):
            # Проверяем различные поля в объекте
            for key in ("checked", "checkmark", "value", "text"):
                val = value.get(key)
                if isinstance(val, bool):
                    return val
                if isinstance(val, str):
                    return val.lower() in ("да", "yes", "true", "1", "checked")
        
        return False
    
    def _get_month_value(self, task_fields: List[Dict[str, Any]], field_id: int) -> str:
        """Получает значение месяца из справочника (поле 181 для формы 792300)."""
        value = self._get_field_value(task_fields, field_id)
        
        if isinstance(value, dict):
            # Проверяем choice_names для справочника выбора
            choice_names = value.get("choice_names")
            if isinstance(choice_names, list) and len(choice_names) > 0:
                return str(choice_names[0])
            
            # Проверяем массив values
            values = value.get("values")
            if isinstance(values, list) and len(values) > 0:
                return str(values[0])
            
            # Проверяем rows
            rows = value.get("rows")
            if isinstance(rows, list) and len(rows) > 0 and isinstance(rows[0], list) and len(rows[0]) > 0:
                return str(rows[0][0])
            
            # Проверяем text, name, value
            for key in ("text", "name", "value"):
                month_val = value.get(key)
                if isinstance(month_val, str) and month_val.strip():
                    return month_val.strip()
        
        if isinstance(value, str):
            return value.strip()
        
        return ""
    
    def _is_studying_september(self, month_value: str) -> bool:
        """Проверяет, является ли месяц 'Сентябрь' (для формы 792300)."""
        return month_value.lower() in ("сентябрь", "september")
    
    def _parse_date_value(self, value: Any) -> Optional[datetime]:
        """Парсит значение даты из различных форматов Pyrus API."""
        if value is None:
            return None
        
        # Строковое значение даты
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            # Пробуем различные форматы
            for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%fZ"]:
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue
        
        # Объект с датой
        if isinstance(value, dict):
            # ISO формат в поле date
            date_str = value.get("date")
            if isinstance(date_str, str):
                try:
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except ValueError:
                    pass
            
            # Текстовое представление
            text_val = value.get("text") or value.get("value")
            if isinstance(text_val, str):
                return self._parse_date_value(text_val)
        
        return None
    
    def _is_date_in_august_september_2025(self, date_value: Any) -> bool:
        """Проверяет, попадает ли дата в диапазон август-сентябрь 2025."""
        # Если date_value уже datetime объект - используем напрямую
        if isinstance(date_value, datetime):
            parsed_date = date_value
        else:
            # Иначе пытаемся распарсить
            parsed_date = self._parse_date_value(date_value)
        
        if parsed_date is None:
            return False
        
        # Диапазон: 01.08.2025 - 30.09.2025
        start_date = datetime(2025, 8, 1)
        end_date = datetime(2025, 9, 30, 23, 59, 59)
        
        return start_date <= parsed_date <= end_date
    
    def _validate_dates_form_2304918(self, task_fields: List[Dict[str, Any]]) -> bool:
        """
        Валидирует даты для формы 2304918 (поля 26, 31, 56).
        
        Правила:
        - Если ВСЕ поля пустые → ВАЛИДНО (студент не вышел на обучение)
        - Если хотя бы одно поле заполнено → ВСЕ заполненные даты должны быть из августа-сентября 2025
        
        Returns:
            True если форма проходит валидацию, False иначе
        """
        date_field_ids = [26, 31, 56]
        
        found_dates = []
        valid_dates = []
        
        for field_id in date_field_ids:
            value = self._get_field_value(task_fields, field_id)
            
            # Пропускаем пустые поля
            if value is None:
                continue
            
            # Пробуем распарсить дату
            parsed_date = self._parse_date_value(value)
            
            if parsed_date is not None:
                found_dates.append(parsed_date)
                
                # Проверяем, попадает ли дата в диапазон август-сентябрь 2025
                if self._is_date_in_august_september_2025(parsed_date):
                    valid_dates.append(parsed_date)
        
        # Если нет ни одной даты - ВАЛИДНО (студент не вышел на обучение)
        if len(found_dates) == 0:
            return True
        
        # Если не все даты из августа-сентября 2025 - форма не проходит
        if len(valid_dates) != len(found_dates):
            return False
        
        # Есть хотя бы одна дата и все даты валидные
        return True
    
    def _validate_date_form_792300(self, task_fields: List[Dict[str, Any]]) -> bool:
        """
        Валидирует дату БПЗ для формы 792300 (поле 220).
        
        Правила:
        - Поле 220 (дата БПЗ) обязательно и должно быть из августа-сентября 2025
        
        Returns:
            True если форма проходит валидацию, False иначе
        """
        date_field_id = 220  # ИЗМЕНЕНО: было 197, стало 220 (дата БПЗ)
        
        value = self._get_field_value(task_fields, date_field_id)
        
        # Поле обязательно - если пустое, форма не проходит
        if value is None:
            return False
        
        # Дата должна быть из августа-сентября 2025
        return self._is_date_in_august_september_2025(value)
    
    async def analyze_form_2304918(self) -> None:
        """Анализ формы 2304918 (возврат студентов) с ПОЛНОЙ отладкой."""
        print("Анализ формы 2304918 (старички)...")
        
        form_id = 2304918
        excluded_count = 0  # Счетчик исключенных преподавателей
        teacher_field_id = 8  # Поле с преподавателем
        studying_field_id = 64  # Поле "УЧИТСЯ (заполняет СО)"
        branch_field_id = 5  # Поле с филиалом
        status_field_id = 7  # Поле со статусом PE
        student_field_id = 2  # Поле с ФИО студента
        
        task_count = 0
        filtered_count = 0
        
        # КРИТИЧЕСКИ ВАЖНО: создаем отдельный счетчик для каждого преподавателя
        teacher_counters = defaultdict(int)
        
        async for task in self.client.iter_register_tasks(form_id, include_archived=True):
            task_count += 1
            if task_count % 100 == 0:
                print(f"Обработано {task_count} задач формы 2304918...")
            
            task_fields = task.get("fields", [])
            task_id = task.get("id")
            
            # СНАЧАЛА проверяем PE статус (самый ранний фильтр)
            if not self._is_valid_pe_status(task_fields, status_field_id):
                continue  # Просто пропускаем, не добавляем в статистику
            
            # Извлекаем преподавателя ПОСЛЕ проверки PE
            teacher_name = self._extract_teacher_name(task_fields, teacher_field_id)
            
            # ОТЛАДКА: считаем ВСЕ найденные задачи для целевого преподавателя
            if teacher_name == self.debug_target:
                self.debug_counters["2304918_found"] += 1
            
            # ОТЛАДКА: считаем задачи с валидным PE для целевого преподавателя
            if teacher_name == self.debug_target:
                self.debug_counters["2304918_valid_pe"] += 1
            
            # Проверяем даты в полях 26, 31, 56 (август-сентябрь 2025)
            if not self._validate_dates_form_2304918(task_fields):
                continue
            
            # ОТЛАДКА: считаем задачи с валидными датами для целевого преподавателя
            if teacher_name == self.debug_target:
                self.debug_counters["2304918_valid_dates"] += 1
            
            filtered_count += 1
            
            # Извлекаем филиал
            branch_name = self._extract_branch_name(task_fields, branch_field_id)
            
            # Извлекаем ФИО студента
            student_name = self._extract_teacher_name(task_fields, student_field_id)
            
            # Проверяем отметку "учится"
            is_studying = self._is_studying(task_fields, studying_field_id)
            
            # Получаем статус PE для проверки (нужно для исключения "Китайского" из филиалов)
            pe_status_value = self._get_field_value(task_fields, status_field_id)
            is_chinese = False
            if isinstance(pe_status_value, dict):
                choice_names = pe_status_value.get("choice_names", [])
                if isinstance(choice_names, list) and len(choice_names) > 0:
                    is_chinese = "китайский" in str(choice_names[0]).lower()
            
            # Учитываем в статистике филиала ТОЛЬКО если:
            # 1) филиал НЕ исключен из соревнования
            # 2) статус PE НЕ "Китайский" (для формы 2304918)
            if not self._is_branch_excluded_from_competition(branch_name) and not is_chinese:
                # Инициализируем статистику филиала если нужно
                if branch_name not in self.branches_stats:
                    self.branches_stats[branch_name] = BranchStats(branch_name)
                
                branch_stats = self.branches_stats[branch_name]
                branch_stats.form_2304918_total += 1
                if is_studying:
                    branch_stats.form_2304918_studying += 1
            
            # Проверяем исключения для старичков (форма 2304918) - только для статистики преподавателей
            if self._is_teacher_excluded(teacher_name, 'oldies'):
                excluded_count += 1
                
                # ОТЛАДКА: считаем исключенные задачи для целевого преподавателя
                if teacher_name == self.debug_target:
                    self.debug_counters["2304918_excluded"] += 1
                
                continue  # Не добавляем в статистику преподавателей
            
            # КРИТИЧЕСКИ ВАЖНО: инициализируем статистику преподавателя только ОДИН раз
            if teacher_name not in self.teachers_stats:
                self.teachers_stats[teacher_name] = TeacherStats(teacher_name)
                
                # ОТЛАДКА: логируем создание нового преподавателя
                if teacher_name == self.debug_target:
                    print(f"   🆕 СОЗДАН новый преподаватель: {teacher_name}")
            
            teacher_stats = self.teachers_stats[teacher_name]
            
            # КРИТИЧЕСКИ ВАЖНО: увеличиваем счетчики АТОМАРНО
            teacher_stats.form_2304918_total += 1
            if is_studying:
                teacher_stats.form_2304918_studying += 1
            
            # Увеличиваем отладочный счетчик
            teacher_counters[teacher_name] += 1
            
            # ОТЛАДКА: считаем обработанные задачи для целевого преподавателя
            if teacher_name == self.debug_target:
                self.debug_counters["2304918_processed"] += 1
                print(f"   🔄 ОБРАБОТАНО {self.debug_counters['2304918_processed']}: {teacher_name} → итого {teacher_stats.form_2304918_total}, учится {teacher_stats.form_2304918_studying}")
            
            # Сохраняем данные для детализации
            teacher_stats.form_2304918_data.append({
                "task_id": task_id,
                "teacher": teacher_name,
                "branch": branch_name,
                "student_name": student_name,
                "is_studying": is_studying
            })
        
        print(f"Завершен анализ формы 2304918. Обработано {task_count} задач, отфильтровано {filtered_count} с валидным статусом PE, исключено {excluded_count} преподавателей.")
        
        # ОТЛАДКА: проверяем финальное состояние для целевого преподавателя
        if self.debug_target in self.teachers_stats:
            final_stats = self.teachers_stats[self.debug_target]
            print(f"   🎯 ФИНАЛЬНОЕ СОСТОЯНИЕ {self.debug_target}: {final_stats.form_2304918_total} всего, {final_stats.form_2304918_studying} учится")
        else:
            print(f"   ❌ {self.debug_target} НЕ НАЙДЕН в финальной статистике!")
    
    async def analyze_form_792300(self) -> None:
        """Анализ формы 792300 (конверсия после БПЗ) с новой логикой поля 183."""
        print("Анализ формы 792300 (новый клиент) с полем 183...")
        
        form_id = 792300
        excluded_count = 0  # Счетчик исключенных преподавателей
        teacher_field_id = 142  # Поле с преподавателем
        month_field_id = 181  # Поле с месяцем
        branch_field_id = 226  # Поле с филиалом
        status_field_id = 228  # Поле со статусом PE
        student_field_id = 73  # Поле с ФИО студента
        field_183_id = 183  # Поле "пришёл на БПЗ"
        
        task_count = 0
        
        # Словарь для группировки форм по клиентам: {student_key: [формы]}
        students_forms = defaultdict(list)
        
        # КРИТИЧЕСКИ ВАЖНО: создаем отдельный счетчик для каждого преподавателя
        teacher_counters = defaultdict(int)
        
        async for task in self.client.iter_register_tasks(form_id, include_archived=True):
            task_count += 1
            if task_count % 100 == 0:
                print(f"Обработано {task_count} задач формы 792300...")
            
            task_fields = task.get("fields", [])
            task_id = task.get("id")
            
            # СНАЧАЛА проверяем PE статус (самый ранний фильтр)
            if not self._is_valid_pe_status(task_fields, status_field_id):
                continue  # Просто пропускаем, не добавляем в статистику
            
            # Извлекаем данные формы
            teacher_name = self._extract_teacher_name(task_fields, teacher_field_id)
            branch_name = self._extract_branch_name(task_fields, branch_field_id)
            student_name = self._extract_teacher_name(task_fields, student_field_id)
            date_value = self._get_field_value(task_fields, 220)  # Дата БПЗ
            month_value = self._get_month_value(task_fields, month_field_id)
            field_183_value = self._get_month_value(task_fields, field_183_id)
            
            # ОТЛАДКА: считаем ВСЕ найденные задачи для целевого преподавателя
            if teacher_name == self.debug_target:
                self.debug_counters["792300_found"] += 1
            
            # Группируем формы по клиенту
            student_key = f"{student_name}|{branch_name}"
            
            students_forms[student_key].append({
                "task_id": task_id,
                "teacher_name": teacher_name,
                "branch_name": branch_name,
                "student_name": student_name,
                "date_bpz_obj": self._parse_date_value(date_value),
                "month": month_value or "",
                "field_183": field_183_value or "",
            })
        
        print(f"✅ Загрузка завершена. Найдено {len(students_forms)} уникальных клиентов из {task_count} форм.")
        
        # Теперь анализируем каждого клиента
        print("🔍 Анализирую клиентов по новой логике...")
        
        filtered_count = 0
        
        for student_key, forms in students_forms.items():
            # БАЗА: Есть ли хотя бы одна форма с БПЗ в августе-сентябре + поле 183 = "Да"?
            has_valid_bpz = False
            valid_forms = []
            
            for form in forms:
                is_date_valid = form["date_bpz_obj"] and self._is_date_in_august_september_2025(form["date_bpz_obj"])
                is_183_yes = form["field_183"].lower() in ("да", "yes")
                
                if is_date_valid and is_183_yes:
                    has_valid_bpz = True
                    valid_forms.append(form)
            
            # Если клиент не попал в базу - пропускаем
            if not has_valid_bpz:
                continue
            
            # УЧИТСЯ: Есть ли хотя бы одна форма с 183 = "Да" И месяцем "Сентябрь"?
            has_september = False
            
            for form in forms:
                is_183_yes = form["field_183"].lower() in ("да", "yes")
                is_month_september = form["month"].lower() in ("сентябрь", "september")
                
                if is_183_yes and is_month_september:
                    has_september = True
                    break
            
            # Берём первую валидную форму для извлечения данных
            display_form = valid_forms[0]
            teacher_name = display_form["teacher_name"]
            branch_name = display_form["branch_name"]
            student_name = display_form["student_name"]
            
            filtered_count += 1
            
            # ОТЛАДКА: считаем задачи с валидными данными для целевого преподавателя
            if teacher_name == self.debug_target:
                self.debug_counters["792300_valid_dates"] += 1
            
            # Учитываем в статистике филиала ТОЛЬКО если филиал НЕ исключен из соревнования
            if not self._is_branch_excluded_from_competition(branch_name):
                # Инициализируем статистику филиала если нужно
                if branch_name not in self.branches_stats:
                    self.branches_stats[branch_name] = BranchStats(branch_name)
                
                branch_stats = self.branches_stats[branch_name]
                branch_stats.form_792300_total += 1
                if has_september:
                    branch_stats.form_792300_studying += 1
            
            # Проверяем исключения для БПЗ (форма 792300) - только для статистики преподавателей
            if self._is_teacher_excluded(teacher_name, 'trial'):
                excluded_count += 1
                
                # ОТЛАДКА: считаем исключенные задачи для целевого преподавателя
                if teacher_name == self.debug_target:
                    self.debug_counters["792300_excluded"] += 1
                
                continue  # Не добавляем в статистику преподавателей
            
            # КРИТИЧЕСКИ ВАЖНО: инициализируем статистику преподавателя только ОДИН раз
            if teacher_name not in self.teachers_stats:
                self.teachers_stats[teacher_name] = TeacherStats(teacher_name)
                
                # ОТЛАДКА: логируем создание нового преподавателя
                if teacher_name == self.debug_target:
                    print(f"   🆕 СОЗДАН новый преподаватель в 792300: {teacher_name}")
            
            teacher_stats = self.teachers_stats[teacher_name]
            
            # КРИТИЧЕСКИ ВАЖНО: увеличиваем счетчики АТОМАРНО
            teacher_stats.form_792300_total += 1
            if has_september:
                teacher_stats.form_792300_studying += 1
            
            # Увеличиваем отладочный счетчик
            teacher_counters[teacher_name] += 1
            
            # ОТЛАДКА: считаем обработанные задачи для целевого преподавателя
            if teacher_name == self.debug_target:
                self.debug_counters["792300_processed"] += 1
                print(f"   🔄 ОБРАБОТАНО {self.debug_counters['792300_processed']}: {teacher_name} → итого 792300: {teacher_stats.form_792300_total}, учится {teacher_stats.form_792300_studying}")
            
            # Сохраняем данные для детализации
            teacher_stats.form_792300_data.append({
                "task_id": display_form["task_id"],
                "teacher": teacher_name,
                "branch": branch_name,
                "student_name": student_name,
                "is_studying": has_september
            })
        
        print(f"Завершен анализ формы 792300. Обработано {task_count} задач, отфильтровано {filtered_count} клиентов в базе, исключено {excluded_count} преподавателей.")
        
        # ОТЛАДКА: проверяем финальное состояние для целевого преподавателя
        if self.debug_target in self.teachers_stats:
            final_stats = self.teachers_stats[self.debug_target]
            print(f"   🎯 ФИНАЛЬНОЕ СОСТОЯНИЕ {self.debug_target}: 2304918={final_stats.form_2304918_total}, 792300={final_stats.form_792300_total}")
        else:
            print(f"   ❌ {self.debug_target} НЕ НАЙДЕН в финальной статистике!")
    
    def create_excel_reports(self, filename: str = "final_teacher_report.xlsx") -> None:
        """Создает полный Excel файл с 3 вкладками: Вывод старичков, Конверсия после БПЗ, Статистика по филиалам."""
        print(f"Создание ОКОНЧАТЕЛЬНО ИСПРАВЛЕННОГО Excel отчета: {filename}")
        
        # Создаем один файл с тремя листами
        wb = Workbook()
        
        # Удаляем дефолтный лист
        wb.remove(wb.active)
        
        # Вкладка 1: Вывод старичков (форма 2304918)
        print("Создание вкладки 'Вывод старичков'...")
        self._create_oldies_sheet(wb)
        
        # Вкладка 2: Конверсия после БПЗ (форма 792300)
        print("Создание вкладки 'Конверсия после БПЗ'...")
        self._create_trial_sheet(wb)
        
        # Вкладка 3: Статистика по филиалам
        if self.branches_stats:
            print(f"Создание вкладки 'Статистика по филиалам': {len(self.branches_stats)} филиалов")
            self._create_branch_summary_sheet(wb)
        else:
            print("⚠️ Нет данных по филиалам для создания вкладки")
        
        # Вкладки 4+: Детальные вкладки по каждому филиалу
        print("Создание детальных вкладок по филиалам...")
        self._create_branch_detail_sheets(wb)
        
        # Сохраняем файл
        wb.save(filename)
        print(f"✅ ОКОНЧАТЕЛЬНО ИСПРАВЛЕННЫЙ полный отчет сохранен: {filename}")
        
        # Подсчитываем общее количество вкладок
        total_sheets = len(wb.sheetnames)
        branch_detail_sheets = total_sheets - 3  # Основные 3 вкладки
        print(f"Файл содержит {total_sheets} вкладок:")
        print("  📊 3 основные: Вывод старичков, Конверсия после БПЗ, Статистика по филиалам")
        print(f"  🏢 {branch_detail_sheets} детальных по филиалам")
    
    def _create_oldies_sheet(self, wb: Workbook) -> None:
        """Создает вкладку 'Вывод старичков' с группировкой по количеству студентов и призами."""
        ws = wb.create_sheet("Вывод старичков")
        
        # Добавляем правила формирования таблицы
        rules_text = "Учитываются формы 2304918 со статусом PE: Start, Future, PE 5, Китайский. Даты выхода (поля 26,31,56): если пусто - включаем, если заполнено - только август-сентябрь 2025. Процент = доля форм со статусом 'учится'."
        ws.cell(row=1, column=1, value=rules_text)
        ws.cell(row=1, column=1).font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A1:E1')  # Объединяем ячейки для правил
        
        # Заголовки (теперь во второй строке)
        headers = [
            "👨‍🏫 Преподаватель",
            "📊 Всего",
            "🎓 Учится", 
            "📈 %",
            "🏆 Приз"
        ]
        
        # Применяем заголовки (теперь во второй строке)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Группируем преподавателей по количеству студентов (форма 2304918)
        # Порядок: от большего к меньшему
        groups = {
            "35+": [],
            "16-34": [], 
            "6-15": []
        }
        
        for teacher_name, stats in self.teachers_stats.items():
            student_count = stats.form_2304918_total
            if 6 <= student_count <= 15:
                groups["6-15"].append(stats)
            elif 16 <= student_count <= 34:
                groups["16-34"].append(stats)
            elif student_count >= 35:
                groups["35+"].append(stats)
        
        # ОТЛАДКА: показываем где попадает целевой преподаватель
        if self.debug_target in self.teachers_stats:
            target_stats = self.teachers_stats[self.debug_target]
            student_count = target_stats.form_2304918_total
            if 6 <= student_count <= 15:
                group = "6-15"
            elif 16 <= student_count <= 34:
                group = "16-34"
            elif student_count >= 35:
                group = "35+"
            else:
                group = "< 6"
            print(f"   🎯 ГРУППИРОВКА: {self.debug_target} ({student_count} форм) → группа {group}")
        
        # Определяем призы для каждой группы
        prize_configs = {
            "35+": {"prizes": ["iPad", "HonorPad", "HonorPad", "HonorPad"], "count": 4},
            "16-34": {"prize": "HonorPad", "count": 3},
            "6-15": {"prize": "Подписка в Tg Premium", "count": 3}
        }
        
        row = 3
        
        # Обрабатываем каждую группу
        for group_name, teachers_list in groups.items():
            if not teachers_list:
                continue
                
            # Добавляем заголовок группы
            group_emojis = {"35+": "🥇", "16-34": "🥈", "6-15": "🥉"}
            emoji = group_emojis.get(group_name, "📋")
            ws.cell(row=row, column=1, value=f"{emoji} Группа {group_name} студентов:")
            ws.cell(row=row, column=1).font = Font(bold=True, color="0066CC")
            row += 1
            
            # Сортируем по % возврата, при равенстве - по количеству клиентов
            sorted_teachers = sorted(
                teachers_list, 
                key=lambda x: (x.return_percentage, x.form_2304918_total), 
                reverse=True
            )
            
            # Определяем призы
            config = prize_configs[group_name]
            for i, stats in enumerate(sorted_teachers):
                prize = ""
                if i < config["count"]:
                    if group_name == "35+" and "prizes" in config:
                        base_prize = config["prizes"][i]
                        if base_prize == "iPad":
                            prize = "📱 iPad"
                        elif base_prize == "HonorPad":
                            prize = "📲 HonorPad"
                    else:
                        base_prize = config["prize"]
                        if base_prize == "Подписка в Tg Premium":
                            prize = "💎 Подписка в Tg Premium"
                        elif base_prize == "HonorPad":
                            prize = "📲 HonorPad"
                
                ws.cell(row=row, column=1, value=stats.name)
                ws.cell(row=row, column=2, value=stats.form_2304918_total)
                ws.cell(row=row, column=3, value=stats.form_2304918_studying)
                ws.cell(row=row, column=4, value=round(stats.return_percentage, 2))
                ws.cell(row=row, column=5, value=prize)
                
                # ОТЛАДКА: логируем что записываем в Excel для целевого преподавателя
                if stats.name == self.debug_target:
                    print(f"   📝 ЗАПИСЫВАЕМ В EXCEL: {stats.name} → {stats.form_2304918_total} форм, {stats.form_2304918_studying} учится, {stats.return_percentage:.2f}%")
                
                # Выделяем призеров
                if prize:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFD700", end_color="FFD700", fill_type="solid"
                        )
                
                row += 1
            
            # Добавляем пустую строку между группами
            row += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Пропускаем объединенные ячейки
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_trial_sheet(self, wb: Workbook) -> None:
        """Создает вкладку 'Конверсия после БПЗ' с группировкой по количеству БПЗ студентов и призами."""
        ws = wb.create_sheet("Конверсия после БПЗ")
        
        # Добавляем правила формирования таблицы
        rules_text = "Учитываются формы 792300 со статусом PE: Start, Future, PE 5, Китайский. Дата выхода (поле 197): если пусто - включаем, если заполнено - только август-сентябрь 2025. Процент = доля форм со статусом 'учится'."
        ws.cell(row=1, column=1, value=rules_text)
        ws.cell(row=1, column=1).font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A1:E1')  # Объединяем ячейки для правил
        
        # Заголовки (теперь во второй строке)
        headers = [
            "👨‍🏫 Преподаватель",
            "📊 Всего",
            "🎓 Учится",
            "📈 %",
            "🏆 Приз"
        ]
        
        # Применяем заголовки (теперь во второй строке)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Группируем преподавателей по количеству БПЗ студентов (форма 792300)
        # Порядок: от большего к меньшему
        groups = {
            "16+": [],
            "11-15": [],
            "5-10": []
        }
        
        for teacher_name, stats in self.teachers_stats.items():
            bpz_count = stats.form_792300_total
            if 5 <= bpz_count <= 10:
                groups["5-10"].append(stats)
            elif 11 <= bpz_count <= 15:
                groups["11-15"].append(stats)
            elif bpz_count >= 16:
                groups["16+"].append(stats)
        
        # ОТЛАДКА: показываем где попадает целевой преподаватель в БПЗ
        if self.debug_target in self.teachers_stats:
            target_stats = self.teachers_stats[self.debug_target]
            bpz_count = target_stats.form_792300_total
            if 5 <= bpz_count <= 10:
                group = "5-10"
            elif 11 <= bpz_count <= 15:
                group = "11-15"
            elif bpz_count >= 16:
                group = "16+"
            else:
                group = "< 5"
            print(f"   🎯 ГРУППИРОВКА БПЗ: {self.debug_target} ({bpz_count} форм) → группа {group}")
        
        # Определяем призы для каждой группы
        prize_configs = {
            "16+": {"prizes": ["iPad", "HonorPad", "HonorPad", "HonorPad"], "count": 4},
            "11-15": {"prize": "HonorPad", "count": 3},
            "5-10": {"prize": "Подписка в Tg Premium", "count": 3}
        }
        
        row = 3
        
        # Обрабатываем каждую группу
        for group_name, teachers_list in groups.items():
            if not teachers_list:
                continue
                
            # Добавляем заголовок группы
            group_emojis = {"16+": "🥇", "11-15": "🥈", "5-10": "🥉"}
            emoji = group_emojis.get(group_name, "📋")
            ws.cell(row=row, column=1, value=f"{emoji} Группа {group_name} БПЗ студентов:")
            ws.cell(row=row, column=1).font = Font(bold=True, color="0066CC")
            row += 1
            
            # Сортируем по % конверсии, при равенстве - по количеству БПЗ студентов
            sorted_teachers = sorted(
                teachers_list,
                key=lambda x: (x.conversion_percentage, x.form_792300_total),
                reverse=True
            )
            
            # Определяем призы
            config = prize_configs[group_name]
            for i, stats in enumerate(sorted_teachers):
                prize = ""
                if i < config["count"]:
                    if group_name == "16+" and "prizes" in config:
                        base_prize = config["prizes"][i]
                        if base_prize == "iPad":
                            prize = "📱 iPad"
                        elif base_prize == "HonorPad":
                            prize = "📲 HonorPad"
                    else:
                        base_prize = config["prize"]
                        if base_prize == "Подписка в Tg Premium":
                            prize = "💎 Подписка в Tg Premium"
                        elif base_prize == "HonorPad":
                            prize = "📲 HonorPad"
                
                ws.cell(row=row, column=1, value=stats.name)
                ws.cell(row=row, column=2, value=stats.form_792300_total)
                ws.cell(row=row, column=3, value=stats.form_792300_studying)
                ws.cell(row=row, column=4, value=round(stats.conversion_percentage, 2))
                ws.cell(row=row, column=5, value=prize)
                
                # ОТЛАДКА: логируем что записываем в Excel для целевого преподавателя
                if stats.name == self.debug_target:
                    print(f"   📝 ЗАПИСЫВАЕМ В EXCEL БПЗ: {stats.name} → {stats.form_792300_total} форм, {stats.form_792300_studying} учится, {stats.conversion_percentage:.2f}%")
                
                # Выделяем призеров
                if prize:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFD700", end_color="FFD700", fill_type="solid"
                        )
                
                row += 1
            
            # Добавляем пустую строку между группами
            row += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Пропускаем объединенные ячейки
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_branch_summary_sheet(self, wb: Workbook) -> None:
        """Создает лист со статистикой по филиалам."""
        ws = wb.create_sheet("Статистика по филиалам")
        
        # Добавляем правила формирования таблицы
        rules_text = "Суммарная статистика по филиалам (статус PE: Start, Future, PE 5, Китайский). Даты выхода: пустые включаются, заполненные - только август-сентябрь 2025. Итоговый % = % возврата старичков + % конверсии после БПЗ."
        ws.cell(row=1, column=1, value=rules_text)
        ws.cell(row=1, column=1).font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A1:I1')  # Объединяем ячейки для правил
        
        # Заголовки (теперь во второй строке)
        headers = [
            "🏢 Филиал",
            "👴 Ст: Всего",
            "🎓 Ст: Учится",
            "📊 Ст: %",
            "👶 Нов.: Всего",
            "🎓 Нов.: Учится", 
            "📊 Нов.: %",
            "🏆 Итого %",
            "🎁 Приз"
        ]
        
        # Применяем заголовки (теперь во второй строке)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Сортируем филиалы по итоговому проценту (по убыванию)
        sorted_branches = sorted(
            self.branches_stats.values(),
            key=lambda x: x.total_percentage,
            reverse=True
        )
        
        # Определяем призы для топ-5 филиалов
        branch_prizes = [
            "📱 Interactive Display",      # 1 место
            "☕ Кофемашина + 2 кг кофе",   # 2 место  
            "☕ Кофемашина",               # 3 место
            "💰 20 000 руб.",             # 4 место
            "💰 10 000 руб."              # 5 место
        ]
        
        # Данные по филиалам
        row = 3
        for i, branch_stats in enumerate(sorted_branches):
            # Определяем приз
            prize = ""
            if i < len(branch_prizes):
                prize = branch_prizes[i]
            
            ws.cell(row=row, column=1, value=branch_stats.name)
            ws.cell(row=row, column=2, value=branch_stats.form_2304918_total)
            ws.cell(row=row, column=3, value=branch_stats.form_2304918_studying)
            ws.cell(row=row, column=4, value=round(branch_stats.return_percentage, 2))
            ws.cell(row=row, column=5, value=branch_stats.form_792300_total)
            ws.cell(row=row, column=6, value=branch_stats.form_792300_studying)
            ws.cell(row=row, column=7, value=round(branch_stats.conversion_percentage, 2))
            ws.cell(row=row, column=8, value=round(branch_stats.total_percentage, 2))
            ws.cell(row=row, column=9, value=prize)
            
            # Выделяем призеров ярким желтым
            if prize:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFD700", end_color="FFD700", fill_type="solid"
                    )
            
            row += 1
        
        # Добавляем список исключенных филиалов
        row += 2  # Пропускаем строку
        ws.cell(row=row, column=1, value="Филиалы, исключенные из соревнования:")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="CC0000")
        row += 1
        
        excluded_branches = [
            "• Макеева 15 (исключен из соревнования)",
            "• Коммуны 106/1 (исключен из соревнования)",
            "• Славы 30 (исключен из соревнования)", 
            "• Online (исключен из соревнования)"
        ]
        
        for excluded_branch in excluded_branches:
            ws.cell(row=row, column=1, value=excluded_branch)
            ws.cell(row=row, column=1).font = Font(italic=True, color="999999")
            row += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Пропускаем объединенные ячейки
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_branch_detail_sheets(self, wb: Workbook) -> None:
        """Создает детальные вкладки для каждого филиала с группировкой по преподавателям."""
        print("Создание детальных вкладок по филиалам...")
        
        # Собираем все филиалы из данных преподавателей
        all_branches = set()
        for teacher_stats in self.teachers_stats.values():
            for data in teacher_stats.form_2304918_data:
                all_branches.add(data["branch"])
            for data in teacher_stats.form_792300_data:
                all_branches.add(data["branch"])
        
        if not all_branches:
            print("⚠️ Нет данных для создания вкладок филиалов")
            return
        
        created_sheets = 0
        
        # Создаем вкладку для каждого филиала
        for branch_name in sorted(all_branches):
            # Проверяем, есть ли преподаватели с данными в этом филиале
            teachers_in_branch = self._get_teachers_in_branch(branch_name)
            
            if teachers_in_branch:
                self._create_single_branch_sheet(wb, branch_name, {})
                created_sheets += 1
                print(f"   ✅ Создана вкладка: {branch_name}")
            else:
                print(f"   ⏭️ Пропущен филиал без данных: {branch_name}")
        
        print(f"Создано {created_sheets} вкладок филиалов")

    def _make_safe_sheet_name(self, branch_name: str) -> str:
        """Создает безопасное название листа для Excel, удаляя недопустимые символы."""
        # Excel не разрешает следующие символы в названиях листов: : \ / ? * [ ]
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        
        safe_name = branch_name
        for char in invalid_chars:
            safe_name = safe_name.replace(char, '_')
        
        # Обрезаем до максимальной длины (31 символ для Excel)
        if len(safe_name) > 31:
            safe_name = safe_name[:31]
        
        # Убираем пробелы в начале и конце
        safe_name = safe_name.strip()
        
        # Если название пустое, используем запасное
        if not safe_name:
            safe_name = "Филиал"
        
        return safe_name

    def _create_single_branch_sheet(self, wb: Workbook, branch_name: str, branch_data: Dict[str, Dict[str, Dict[str, int]]]) -> None:
        """Создает отдельную вкладку для филиала с группировкой по преподавателям."""
        # Создаем безопасное название для Excel (удаляем недопустимые символы)
        safe_name = self._make_safe_sheet_name(branch_name)
        ws = wb.create_sheet(safe_name)
        
        current_row = 1
        
        # Заголовок филиала
        ws.cell(row=current_row, column=1, value=f"🏢 ФИЛИАЛ: {branch_name}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="0066CC")
        ws.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 2
        
        # Получаем список всех преподавателей в этом филиале
        teachers_in_branch = self._get_teachers_in_branch(branch_name)
        
        if not teachers_in_branch:
            ws.cell(row=current_row, column=1, value="Нет данных по преподавателям в этом филиале")
            ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
            self._adjust_column_widths(ws)
            return
        
        # Для каждого преподавателя выводим обе секции (старички + БПЗ)
        for i, teacher_info in enumerate(teachers_in_branch):
            current_row = self._add_teacher_section_to_branch_sheet(ws, current_row, teacher_info, branch_name, i)
            current_row += 2  # Разделитель между преподавателями
        
        # Автоширина колонок
        self._adjust_column_widths(ws)

    def _get_teachers_in_branch(self, branch_name: str) -> List[Dict[str, Any]]:
        """Получает список всех преподавателей с данными в указанном филиале."""
        teachers_dict = {}
        
        # Собираем данные по старичкам
        for teacher_name, teacher_stats in self.teachers_stats.items():
            students_oldies = [d for d in teacher_stats.form_2304918_data if d["branch"] == branch_name]
            if students_oldies:
                if teacher_name not in teachers_dict:
                    teachers_dict[teacher_name] = {"name": teacher_name, "oldies": [], "trial": []}
                teachers_dict[teacher_name]["oldies"] = students_oldies
        
        # Собираем данные по БПЗ
        for teacher_name, teacher_stats in self.teachers_stats.items():
            students_trial = [d for d in teacher_stats.form_792300_data if d["branch"] == branch_name]
            if students_trial:
                if teacher_name not in teachers_dict:
                    teachers_dict[teacher_name] = {"name": teacher_name, "oldies": [], "trial": []}
                teachers_dict[teacher_name]["trial"] = students_trial
        
        # Преобразуем в список и добавляем статистику
        teachers_list = []
        for teacher_name, data in teachers_dict.items():
            oldies_total = len(data["oldies"])
            oldies_studying = sum(1 for s in data["oldies"] if s["is_studying"])
            oldies_percentage = (oldies_studying / oldies_total * 100) if oldies_total > 0 else 0
            
            trial_total = len(data["trial"])
            trial_studying = sum(1 for s in data["trial"] if s["is_studying"])
            trial_percentage = (trial_studying / trial_total * 100) if trial_total > 0 else 0
            
            # Общий процент для сортировки
            total_percentage = oldies_percentage + trial_percentage
            
            teachers_list.append({
                "name": teacher_name,
                "oldies_students": data["oldies"],
                "oldies_total": oldies_total,
                "oldies_studying": oldies_studying,
                "oldies_percentage": oldies_percentage,
                "trial_students": data["trial"],
                "trial_total": trial_total,
                "trial_studying": trial_studying,
                "trial_percentage": trial_percentage,
                "total_percentage": total_percentage
            })
        
        # Сортируем по общему проценту (убывание)
        teachers_list.sort(key=lambda x: x["total_percentage"], reverse=True)
        
        return teachers_list

    def _add_teacher_section_to_branch_sheet(self, ws, start_row: int, teacher_info: Dict[str, Any], branch_name: str, teacher_index: int) -> int:
        """Добавляет секцию для одного преподавателя с обеими таблицами (старички + БПЗ)."""
        current_row = start_row
        
        # Заголовок преподавателя (используем простой эмодзи для совместимости с Excel)
        cell = ws.cell(row=current_row, column=1, value=f"👤 Преподаватель: {teacher_info['name']}")
        cell.font = Font(bold=True, size=13)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Выделяем топ-3 преподавателей желтым фоном
        if teacher_index < 3:
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Объединяем ячейки ПОСЛЕ установки всех свойств
        ws.merge_cells(f'A{current_row}:C{current_row}')
        
        current_row += 2
        
        # === СЕКЦИЯ 1: СТАРИЧКИ ===
        if teacher_info["oldies_students"]:
            current_row = self._add_oldies_section(ws, current_row, teacher_info)
            current_row += 1
        
        # === СЕКЦИЯ 2: БПЗ ===
        if teacher_info["trial_students"]:
            current_row = self._add_trial_section(ws, current_row, teacher_info)
            current_row += 1
        
        # Если нет данных вообще
        if not teacher_info["oldies_students"] and not teacher_info["trial_students"]:
            ws.cell(row=current_row, column=1, value="Нет данных по этому преподавателю")
            ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
            current_row += 1
        
        return current_row

    def _add_oldies_section(self, ws, start_row: int, teacher_info: Dict[str, Any]) -> int:
        """Добавляет секцию 'Старички' для преподавателя."""
        current_row = start_row
        
        # Заголовок секции
        ws.cell(row=current_row, column=1, value="📊 СТАРИЧКИ (форма 2304918):")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="0066CC")
        current_row += 1
        
        # Заголовки таблицы
        ws.cell(row=current_row, column=1, value="ФИО студента")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws.cell(row=current_row, column=2, value="Статус")
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
        
        # Разделяем студентов
        studying = [s for s in teacher_info["oldies_students"] if s["is_studying"]]
        not_studying = [s for s in teacher_info["oldies_students"] if not s["is_studying"]]
        
        # Сначала вышедшие
        if studying:
            ws.cell(row=current_row, column=1, value="✅ ВЫШЛИ:")
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="008000")
            current_row += 1
            
            for student in studying:
                ws.cell(row=current_row, column=1, value=student["student_name"])
                ws.cell(row=current_row, column=2, value="✅ Учится")
                current_row += 1
        
        # Потом не вышедшие
        if not_studying:
            ws.cell(row=current_row, column=1, value="❌ НЕ ВЫШЛИ:")
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="CC0000")
            current_row += 1
            
            for student in not_studying:
                ws.cell(row=current_row, column=1, value=student["student_name"])
                ws.cell(row=current_row, column=2, value="❌ Не учится")
                current_row += 1
        
        # Итоговая строка
        ws.cell(row=current_row, column=1, value=f"📊 Итого: {teacher_info['oldies_total']} студентов, вышло {teacher_info['oldies_studying']} ({teacher_info['oldies_percentage']:.1f}%)")
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        return current_row

    def _add_trial_section(self, ws, start_row: int, teacher_info: Dict[str, Any]) -> int:
        """Добавляет секцию 'БПЗ' для преподавателя."""
        current_row = start_row
        
        # Заголовок секции
        ws.cell(row=current_row, column=1, value="📊 КОНВЕРСИЯ ПОСЛЕ БПЗ (форма 792300):")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="0066CC")
        current_row += 1
        
        # Заголовки таблицы
        ws.cell(row=current_row, column=1, value="ФИО студента")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws.cell(row=current_row, column=2, value="Статус")
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        current_row += 1
        
        # Разделяем студентов
        studying = [s for s in teacher_info["trial_students"] if s["is_studying"]]
        not_studying = [s for s in teacher_info["trial_students"] if not s["is_studying"]]
        
        # Сначала остались
        if studying:
            ws.cell(row=current_row, column=1, value="✅ ОСТАЛИСЬ:")
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="008000")
            current_row += 1
            
            for student in studying:
                ws.cell(row=current_row, column=1, value=student["student_name"])
                ws.cell(row=current_row, column=2, value="✅ Учится")
                current_row += 1
        
        # Потом не остались
        if not_studying:
            ws.cell(row=current_row, column=1, value="❌ НЕ ОСТАЛИСЬ:")
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="CC0000")
            current_row += 1
            
            for student in not_studying:
                ws.cell(row=current_row, column=1, value=student["student_name"])
                ws.cell(row=current_row, column=2, value="❌ Не учится")
                current_row += 1
        
        # Итоговая строка
        ws.cell(row=current_row, column=1, value=f"📊 Итого: {teacher_info['trial_total']} БПЗ студентов, остались {teacher_info['trial_studying']} ({teacher_info['trial_percentage']:.1f}%)")
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        return current_row

    def _adjust_column_widths(self, ws) -> None:
        """Настраивает автоширину колонок."""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Пропускаем объединенные ячейки
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

    def print_debug_summary(self) -> None:
        """Выводит итоговую отладочную информацию."""
        print(f"\n" + "=" * 80)
        print(f"🔍 ОТЛАДОЧНАЯ СВОДКА ДЛЯ: {self.debug_target}")
        print("=" * 80)
        print(f"📊 Форма 2304918:")
        print(f"   🔍 Найдено всего: {self.debug_counters['2304918_found']}")
        print(f"   ✅ С валидным PE: {self.debug_counters['2304918_valid_pe']}")
        print(f"   📅 С валидными датами: {self.debug_counters['2304918_valid_dates']}")
        print(f"   ❌ Исключено: {self.debug_counters['2304918_excluded']}")
        print(f"   🔄 Обработано: {self.debug_counters['2304918_processed']}")
        
        print(f"📊 Форма 792300:")
        print(f"   🔍 Найдено всего: {self.debug_counters['792300_found']}")
        print(f"   ✅ С валидным PE: {self.debug_counters['792300_valid_pe']}")
        print(f"   📅 С валидными датами: {self.debug_counters['792300_valid_dates']}")
        print(f"   ❌ Исключено: {self.debug_counters['792300_excluded']}")
        print(f"   🔄 Обработано: {self.debug_counters['792300_processed']}")
        
        if self.debug_target in self.teachers_stats:
            final_stats = self.teachers_stats[self.debug_target]
            print(f"\n🎯 ФИНАЛЬНАЯ СТАТИСТИКА:")
            print(f"   📊 Форма 2304918: {final_stats.form_2304918_total} всего, {final_stats.form_2304918_studying} учится ({final_stats.return_percentage:.2f}%)")
            print(f"   📊 Форма 792300: {final_stats.form_792300_total} всего, {final_stats.form_792300_studying} учится ({final_stats.conversion_percentage:.2f}%)")
            print(f"   🏆 Суммарный процент: {final_stats.total_percentage:.2f}%")
        else:
            print(f"\n❌ {self.debug_target} НЕ НАЙДЕН в финальной статистике!")
        
        print("=" * 80)
    
    async def run_analysis(self) -> None:
        """Запускает полный анализ данных."""
        # Создаем папку для отчетов если не существует
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        
        print("Начинаем создание ОКОНЧАТЕЛЬНО ИСПРАВЛЕННОГО отчета из Pyrus...")
        print(f"Время начала: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Анализируем обе формы
        await self.analyze_form_2304918()
        await self.analyze_form_792300()
        
        # Выводим отладочную сводку
        self.print_debug_summary()
        
        # Выводим краткую статистику
        print("\n=== КРАТКАЯ СТАТИСТИКА ===")
        total_teachers = len(self.teachers_stats)
        print(f"Всего преподавателей: {total_teachers}")
        
        # Создаем полный Excel отчет с категориями по вкладкам
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reports/final_fixed_teacher_report_{timestamp}.xlsx"
        self.create_excel_reports(filename)
        
        print(f"\nОКОНЧАТЕЛЬНО ИСПРАВЛЕННЫЙ анализ завершен: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


async def main():
    """Главная функция скрипта."""
    try:
        analyzer = FinalFixedPyrusDataAnalyzer()
        await analyzer.run_analysis()
    except KeyboardInterrupt:
        print("\nОтмена выполнения пользователем.")
        sys.exit(1)
    except Exception as e:
        print(f"Ошибка выполнения: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    # Проверяем переменные окружения
    if not os.getenv("PYRUS_LOGIN") or not os.getenv("PYRUS_SECURITY_KEY"):
        print("ОШИБКА: Не установлены переменные окружения PYRUS_LOGIN и PYRUS_SECURITY_KEY")
        print("Убедитесь, что .env файл настроен корректно.")
        sys.exit(1)
    
    asyncio.run(main())
