#!/usr/bin/env python3
"""
Скрипт для создания Excel отчета по данным из Pyrus.

Анализирует две формы:
- 2304918: возврат студентов (поле 8 - преподаватель, поле 64 - учится)
- 792300: конверсия trial (поле 142 - преподаватель, поле 187 - учится)

Создает Excel файл с:
- Основной отчет: сводка по преподавателям
- Детализация: разбивка по формам
- Исходные данные: для проверки расчетов
"""

import asyncio
import sys
import os
import json
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Set
from datetime import datetime
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv

# Загружаем переменные окружения из .env файла
load_dotenv()

# Добавляем app в путь для импорта
sys.path.append(str(Path(__file__).parent / "app"))

from pyrus_client import PyrusClient


class TeacherStats:
    """Статистика по преподавателю."""
    
    def __init__(self, name: str):
        self.name = name
        # Форма 2304918 (возврат студентов)
        self.form_2304918_total = 0
        self.form_2304918_studying = 0
        self.form_2304918_data = []  # Исходные данные для проверки
        
        # Форма 792300 (конверсия trial)
        self.form_792300_total = 0
        self.form_792300_studying = 0
        self.form_792300_data = []  # Исходные данные для проверки
    
    @property
    def return_percentage(self) -> float:
        """Процент возврата студентов (форма 2304918)."""
        if self.form_2304918_total == 0:
            return 0.0
        return (self.form_2304918_studying / self.form_2304918_total) * 100
    
    @property
    def conversion_percentage(self) -> float:
        """Процент конверсии trial → студент (форма 792300)."""
        if self.form_792300_total == 0:
            return 0.0
        return (self.form_792300_studying / self.form_792300_total) * 100
    
    @property
    def total_percentage(self) -> float:
        """Суммарный процент двух показателей."""
        return self.return_percentage + self.conversion_percentage


class BranchStats:
    """Статистика по филиалу."""
    
    def __init__(self, name: str):
        self.name = name
        # Форма 2304918 (старички)
        self.form_2304918_total = 0
        self.form_2304918_studying = 0
        
        # Форма 792300 (новый клиент)
        self.form_792300_total = 0
        self.form_792300_studying = 0
    
    @property
    def return_percentage(self) -> float:
        """Процент возврата студентов (старички)."""
        if self.form_2304918_total == 0:
            return 0.0
        return (self.form_2304918_studying / self.form_2304918_total) * 100
    
    @property
    def conversion_percentage(self) -> float:
        """Процент конверсии trial → студент (новый клиент)."""
        if self.form_792300_total == 0:
            return 0.0
        return (self.form_792300_studying / self.form_792300_total) * 100
    
    @property
    def total_percentage(self) -> float:
        """Суммарный процент двух показателей."""
        return self.return_percentage + self.conversion_percentage


class PyrusDataAnalyzer:
    """Анализатор данных из Pyrus для создания Excel отчета."""
    
    def __init__(self):
        self.client = PyrusClient()
        self.teachers_stats: Dict[str, TeacherStats] = {}
        self.branches_stats: Dict[str, BranchStats] = {}
        
        # Загружаем списки исключений преподавателей
        self.excluded_teachers = self._load_exclusions()
    
    def _get_field_value(self, field_list: List[Dict[str, Any]], field_id: int) -> Optional[Any]:
        """Ищет значение поля по id, рекурсивно обходя вложенные секции."""
        for f in field_list or []:
            if f.get("id") == field_id:
                return f.get("value")
            val = f.get("value")
            if isinstance(val, dict) and isinstance(val.get("fields"), list):
                nested = self._get_field_value(val.get("fields") or [], field_id)
                if nested is not None:
                    return nested
        return None
    
    def _extract_teacher_name(self, task_fields: List[Dict[str, Any]], field_id: int) -> str:
        """Извлекает ФИО преподавателя из поля справочника."""
        value = self._get_field_value(task_fields, field_id)
        
        if isinstance(value, dict):
            # Поддержка person-объекта: first_name/last_name
            first_name = value.get("first_name", "")
            last_name = value.get("last_name", "")
            if isinstance(first_name, str) or isinstance(last_name, str):
                full_name = f"{(first_name or '').strip()} {(last_name or '').strip()}".strip()
                if full_name:
                    return full_name
            
            # Для справочника сотрудников обычно есть поле text или name
            for key in ("text", "name", "value"):
                name_val = value.get(key)
                if isinstance(name_val, str) and name_val.strip():
                    return name_val.strip()
        
        if isinstance(value, str):
            return value.strip()
        
        return "Неизвестный преподаватель"
    
    def _load_exclusions(self) -> Dict[str, Set[str]]:
        """Загружает списки исключений преподавателей из JSON файла."""
        exclusions_file = Path(__file__).parent / "teacher_exclusions.json"
        
        try:
            with open(exclusions_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return {
                    'oldies': set(data.get('oldies', [])),
                    'trial': set(data.get('trial', []))
                }
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"⚠️ Не удалось загрузить исключения из {exclusions_file}: {e}")
            print("Продолжаем без фильтрации преподавателей.")
            return {'oldies': set(), 'trial': set()}
    
    def _is_teacher_excluded(self, teacher_name: str, form_type: str) -> bool:
        """Проверяет, исключен ли преподаватель из указанной формы.
        
        Args:
            teacher_name: Имя преподавателя
            form_type: 'oldies' для формы 2304918, 'trial' для формы 792300
        """
        excluded_set = self.excluded_teachers.get(form_type, set())
        
        # Проверяем точное совпадение
        if teacher_name in excluded_set:
            return True
        
        # Проверяем частичное совпадение (фамилия входит в имя преподавателя)
        for excluded_name in excluded_set:
            if excluded_name.lower() in teacher_name.lower():
                return True
        
        return False
    
    def _normalize_branch_name(self, branch_name: str) -> str:
        """Нормализует название филиала для объединения данных."""
        branch_name = branch_name.lower().strip()
        
        # Исключаем филиалы из отчета
        if "макеева" in branch_name and "15" in branch_name:
            return None  # Исключаем из отчета
        if "коммуны" in branch_name and "106/1" in branch_name:
            return None  # Исключаем из отчета
        
        # Объединяем филиалы Копейска под единым названием
        if "коммунистический" in branch_name and "22" in branch_name:
            return "Копейск"
        if "славы" in branch_name and "30" in branch_name:
            return "Копейск"
        
        # Возвращаем оригинальное название с заглавной буквы
        return branch_name.title()
    
    def _extract_branch_name(self, task_fields: List[Dict[str, Any]], field_id: int) -> str:
        """Извлекает название филиала из поля справочника."""
        value = self._get_field_value(task_fields, field_id)
        
        if isinstance(value, dict):
            # Проверяем массив values - основной способ для справочника филиалов
            values = value.get("values")
            if isinstance(values, list) and len(values) > 0:
                branch_name = values[0]  # Берем первое значение
                if isinstance(branch_name, str) and branch_name.strip():
                    return self._normalize_branch_name(branch_name.strip())
            
            # Проверяем rows если values не найден  
            rows = value.get("rows")
            if isinstance(rows, list) and len(rows) > 0 and isinstance(rows[0], list) and len(rows[0]) > 0:
                branch_name = rows[0][0]  # Берем первую ячейку первой строки
                if isinstance(branch_name, str) and branch_name.strip():
                    return self._normalize_branch_name(branch_name.strip())
            
            # Для обычных справочников обычно есть поле text или name
            for key in ("text", "name", "value"):
                branch_val = value.get(key)
                if isinstance(branch_val, str) and branch_val.strip():
                    return self._normalize_branch_name(branch_val.strip())
        
        if isinstance(value, str):
            return self._normalize_branch_name(value.strip())
        
        return "Неизвестный филиал"
    
    def _is_valid_pe_status(self, task_fields: List[Dict[str, Any]], field_id: int) -> bool:
        """Проверяет, соответствует ли статус PE одному из допустимых: PE Start, PE Future, PE 5."""
        value = self._get_field_value(task_fields, field_id)
        
        # Допустимые статусы PE
        valid_statuses = {"PE Start", "PE Future", "PE 5"}
        
        if isinstance(value, dict):
            # Проверяем choice_names для справочника выбора
            choice_names = value.get("choice_names")
            if isinstance(choice_names, list) and len(choice_names) > 0:
                status = choice_names[0]
                if isinstance(status, str) and status.strip() in valid_statuses:
                    return True
            
            # Проверяем массив values для справочника
            values = value.get("values")
            if isinstance(values, list) and len(values) > 0:
                status = values[0]
                if isinstance(status, str) and status.strip() in valid_statuses:
                    return True
            
            # Проверяем rows если values не найден  
            rows = value.get("rows")
            if isinstance(rows, list) and len(rows) > 0 and isinstance(rows[0], list) and len(rows[0]) > 0:
                status = rows[0][0]
                if isinstance(status, str) and status.strip() in valid_statuses:
                    return True
            
            # Для обычных справочников проверяем text, name, value
            for key in ("text", "name", "value"):
                status_val = value.get(key)
                if isinstance(status_val, str) and status_val.strip() in valid_statuses:
                    return True
        
        if isinstance(value, str):
            return value.strip() in valid_statuses
        
        return False
    
    def _is_studying(self, task_fields: List[Dict[str, Any]], field_id: int) -> bool:
        """Проверяет, отмечена ли галочка 'учится' в указанном поле."""
        value = self._get_field_value(task_fields, field_id)
        
        if value is None:
            return False
        
        # Булево значение
        if isinstance(value, bool):
            return value
        
        # Строковое значение (checked/unchecked)
        if isinstance(value, str):
            return value.lower() in ("да", "yes", "true", "checked")
        
        # Объект с чекбоксом
        if isinstance(value, dict):
            checkmark = value.get("checkmark")
            if checkmark == "checked":
                return True
        
        return False
    
    async def analyze_form_2304918(self) -> None:
        """Анализ формы 2304918 (возврат студентов)."""
        print("Анализ формы 2304918 (старички)...")
        
        form_id = 2304918
        excluded_count = 0  # Счетчик исключенных преподавателей
        teacher_field_id = 8  # Поле с преподавателем
        studying_field_id = 64  # Поле "УЧИТСЯ (заполняет СО)"
        branch_field_id = 5  # Поле с филиалом
        status_field_id = 7  # Поле со статусом PE
        
        task_count = 0
        filtered_count = 0
        async for task in self.client.iter_register_tasks(form_id, include_archived=False):
            task_count += 1
            if task_count % 100 == 0:
                print(f"Обработано {task_count} задач формы 2304918...")
            
            task_fields = task.get("fields", [])
            task_id = task.get("id")
            
            # Проверяем статус PE - фильтруем только PE Start, PE Future, PE 5
            if not self._is_valid_pe_status(task_fields, status_field_id):
                continue
            
            filtered_count += 1
            
            # Извлекаем преподавателя
            teacher_name = self._extract_teacher_name(task_fields, teacher_field_id)
            
            # Проверяем исключения для старичков (форма 2304918)
            if self._is_teacher_excluded(teacher_name, 'oldies'):
                excluded_count += 1
                continue
            
            # Извлекаем филиал
            branch_name = self._extract_branch_name(task_fields, branch_field_id)
            
            # Пропускаем исключенные филиалы
            if branch_name is None:
                continue
            
            # Инициализируем статистику преподавателя если нужно
            if teacher_name not in self.teachers_stats:
                self.teachers_stats[teacher_name] = TeacherStats(teacher_name)
            
            # Инициализируем статистику филиала если нужно
            if branch_name not in self.branches_stats:
                self.branches_stats[branch_name] = BranchStats(branch_name)
            
            teacher_stats = self.teachers_stats[teacher_name]
            branch_stats = self.branches_stats[branch_name]
            
            # Увеличиваем общий счетчик
            teacher_stats.form_2304918_total += 1
            branch_stats.form_2304918_total += 1
            
            # Проверяем отметку "учится"
            is_studying = self._is_studying(task_fields, studying_field_id)
            if is_studying:
                teacher_stats.form_2304918_studying += 1
                branch_stats.form_2304918_studying += 1
            
            # Сохраняем данные для детализации
            teacher_stats.form_2304918_data.append({
                "task_id": task_id,
                "teacher": teacher_name,
                "branch": branch_name,
                "is_studying": is_studying
            })
        
        print(f"Завершен анализ формы 2304918. Обработано {task_count} задач, отфильтровано {filtered_count} с валидным статусом PE, исключено {excluded_count} преподавателей.")
    
    async def analyze_form_792300(self) -> None:
        """Анализ формы 792300 (конверсия trial)."""
        print("Анализ формы 792300 (новый клиент)...")
        
        form_id = 792300
        excluded_count = 0  # Счетчик исключенных преподавателей
        teacher_field_id = 142  # Поле с преподавателем
        studying_field_id = 187  # Поле "учится"
        branch_field_id = 226  # Поле с филиалом
        status_field_id = 228  # Поле со статусом PE
        
        task_count = 0
        filtered_count = 0
        async for task in self.client.iter_register_tasks(form_id, include_archived=False):
            task_count += 1
            if task_count % 100 == 0:
                print(f"Обработано {task_count} задач формы 792300...")
            
            task_fields = task.get("fields", [])
            task_id = task.get("id")
            
            # Проверяем статус PE - фильтруем только PE Start, PE Future, PE 5
            if not self._is_valid_pe_status(task_fields, status_field_id):
                continue
            
            filtered_count += 1
            
            # Извлекаем преподавателя
            teacher_name = self._extract_teacher_name(task_fields, teacher_field_id)
            
            # Проверяем исключения для trial (форма 792300)
            if self._is_teacher_excluded(teacher_name, 'trial'):
                excluded_count += 1
                continue
            
            # Извлекаем филиал
            branch_name = self._extract_branch_name(task_fields, branch_field_id)
            
            # Пропускаем исключенные филиалы
            if branch_name is None:
                continue
            
            # Инициализируем статистику преподавателя если нужно
            if teacher_name not in self.teachers_stats:
                self.teachers_stats[teacher_name] = TeacherStats(teacher_name)
            
            # Инициализируем статистику филиала если нужно
            if branch_name not in self.branches_stats:
                self.branches_stats[branch_name] = BranchStats(branch_name)
            
            teacher_stats = self.teachers_stats[teacher_name]
            branch_stats = self.branches_stats[branch_name]
            
            # Увеличиваем общий счетчик
            teacher_stats.form_792300_total += 1
            branch_stats.form_792300_total += 1
            
            # Проверяем отметку "учится"
            is_studying = self._is_studying(task_fields, studying_field_id)
            if is_studying:
                teacher_stats.form_792300_studying += 1
                branch_stats.form_792300_studying += 1
            
            # Сохраняем данные для детализации
            teacher_stats.form_792300_data.append({
                "task_id": task_id,
                "teacher": teacher_name,
                "branch": branch_name,
                "is_studying": is_studying
            })
        
        print(f"Завершен анализ формы 792300. Обработано {task_count} задач, отфильтровано {filtered_count} с валидным статусом PE, исключено {excluded_count} преподавателей.")
    
    def create_excel_reports(self, filename: str = "pyrus_teacher_report.xlsx") -> None:
        """Создает Excel файл с 3 вкладками: Вывод старичков, Конверсия trial, Статистика по филиалам."""
        print(f"Создание Excel отчета: {filename}")
        
        # Создаем один файл с тремя листами
        wb = Workbook()
        
        # Удаляем дефолтный лист
        wb.remove(wb.active)
        
        # Вкладка 1: Вывод старичков (форма 2304918)
        print("Создание вкладки 'Вывод старичков'...")
        self._create_oldies_sheet(wb)
        
        # Вкладка 2: Конверсия trial (форма 792300)
        print("Создание вкладки 'Конверсия trial'...")
        self._create_trial_sheet(wb)
        
        # Вкладка 3: Статистика по филиалам
        if self.branches_stats:
            print(f"Создание вкладки 'Статистика по филиалам': {len(self.branches_stats)} филиалов")
            self._create_branch_summary_sheet(wb)
        else:
            print("⚠️ Нет данных по филиалам для создания вкладки")
        
        # Сохраняем файл
        wb.save(filename)
        print(f"✅ Отчет сохранен: {filename}")
        print("Файл содержит 3 вкладки: Вывод старичков, Конверсия trial, Статистика по филиалам!")
    
    def _create_oldies_sheet(self, wb: Workbook) -> None:
        """Создает вкладку 'Вывод старичков' с группировкой по количеству студентов и призами."""
        ws = wb.create_sheet("Вывод старичков")
        
        # Заголовки
        headers = [
            "👨‍🏫 Преподаватель",
            "📊 Всего",
            "🎓 Учится", 
            "📈 %",
            "🏆 Приз"
        ]
        
        # Применяем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Группируем преподавателей по количеству студентов (форма 2304918)
        # Порядок: от большего к меньшему
        groups = {
            "35+": [],
            "16-34": [], 
            "6-15": []
        }
        
        for teacher_name, stats in self.teachers_stats.items():
            student_count = stats.form_2304918_total
            if 6 <= student_count <= 15:
                groups["6-15"].append(stats)
            elif 16 <= student_count <= 34:
                groups["16-34"].append(stats)
            elif student_count >= 35:
                groups["35+"].append(stats)
        
        # Определяем призы для каждой группы
        prize_configs = {
            "35+": {"prizes": ["iPad", "HonorPad", "HonorPad", "HonorPad"], "count": 4},
            "16-34": {"prize": "HonorPad", "count": 3},
            "6-15": {"prize": "Подписка в Tg Premium", "count": 3}
        }
        
        row = 2
        
        # Обрабатываем каждую группу
        for group_name, teachers_list in groups.items():
            if not teachers_list:
                continue
                
            # Добавляем заголовок группы
            group_emojis = {"35+": "🥇", "16-34": "🥈", "6-15": "🥉"}
            emoji = group_emojis.get(group_name, "📋")
            ws.cell(row=row, column=1, value=f"{emoji} Группа {group_name} студентов:")
            ws.cell(row=row, column=1).font = Font(bold=True, color="0066CC")
            row += 1
            
            # Сортируем по % возврата, при равенстве - по количеству клиентов
            sorted_teachers = sorted(
                teachers_list, 
                key=lambda x: (x.return_percentage, x.form_2304918_total), 
                reverse=True
            )
            
            # Определяем призы
            config = prize_configs[group_name]
            for i, stats in enumerate(sorted_teachers):
                prize = ""
                if i < config["count"]:
                    if group_name == "35+" and "prizes" in config:
                        base_prize = config["prizes"][i]
                        if base_prize == "iPad":
                            prize = "📱 iPad"
                        elif base_prize == "HonorPad":
                            prize = "📲 HonorPad"
                    else:
                        base_prize = config["prize"]
                        if base_prize == "Подписка в Tg Premium":
                            prize = "💎 Подписка в Tg Premium"
                        elif base_prize == "HonorPad":
                            prize = "📲 HonorPad"
                
                ws.cell(row=row, column=1, value=stats.name)
                ws.cell(row=row, column=2, value=stats.form_2304918_total)
                ws.cell(row=row, column=3, value=stats.form_2304918_studying)
                ws.cell(row=row, column=4, value=round(stats.return_percentage, 2))
                ws.cell(row=row, column=5, value=prize)
                
                # Выделяем призеров
                if prize:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFD700", end_color="FFD700", fill_type="solid"
                        )
                
                row += 1
            
            # Добавляем пустую строку между группами
            row += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_trial_sheet(self, wb: Workbook) -> None:
        """Создает вкладку 'Конверсия trial' с группировкой по % конверсии и призами."""
        ws = wb.create_sheet("Конверсия trial")
        
        # Заголовки
        headers = [
            "👨‍🏫 Преподаватель",
            "📊 Всего",
            "🎓 Учится",
            "📈 %",
            "🏆 Приз"
        ]
        
        # Применяем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Группируем преподавателей по количеству БПЗ студентов (форма 792300)
        # Порядок: от большего к меньшему
        groups = {
            "16+": [],
            "11-15": [],
            "5-10": []
        }
        
        for teacher_name, stats in self.teachers_stats.items():
            bpz_count = stats.form_792300_total
            if 5 <= bpz_count <= 10:
                groups["5-10"].append(stats)
            elif 11 <= bpz_count <= 15:
                groups["11-15"].append(stats)
            elif bpz_count >= 16:
                groups["16+"].append(stats)
        
        # Определяем призы для каждой группы
        prize_configs = {
            "16+": {"prizes": ["iPad", "HonorPad", "HonorPad", "HonorPad"], "count": 4},
            "11-15": {"prize": "HonorPad", "count": 3},
            "5-10": {"prize": "Подписка в Tg Premium", "count": 3}
        }
        
        row = 2
        
        # Обрабатываем каждую группу
        for group_name, teachers_list in groups.items():
            if not teachers_list:
                continue
                
            # Добавляем заголовок группы
            group_emojis = {"16+": "🥇", "11-15": "🥈", "5-10": "🥉"}
            emoji = group_emojis.get(group_name, "📋")
            ws.cell(row=row, column=1, value=f"{emoji} Группа {group_name} БПЗ студентов:")
            ws.cell(row=row, column=1).font = Font(bold=True, color="0066CC")
            row += 1
            
            # Сортируем по % конверсии, при равенстве - по количеству БПЗ студентов
            sorted_teachers = sorted(
                teachers_list,
                key=lambda x: (x.conversion_percentage, x.form_792300_total),
                reverse=True
            )
            
            # Определяем призы
            config = prize_configs[group_name]
            for i, stats in enumerate(sorted_teachers):
                prize = ""
                if i < config["count"]:
                    if group_name == "16+" and "prizes" in config:
                        base_prize = config["prizes"][i]
                        if base_prize == "iPad":
                            prize = "📱 iPad"
                        elif base_prize == "HonorPad":
                            prize = "📲 HonorPad"
                    else:
                        base_prize = config["prize"]
                        if base_prize == "Подписка в Tg Premium":
                            prize = "💎 Подписка в Tg Premium"
                        elif base_prize == "HonorPad":
                            prize = "📲 HonorPad"
                
                ws.cell(row=row, column=1, value=stats.name)
                ws.cell(row=row, column=2, value=stats.form_792300_total)
                ws.cell(row=row, column=3, value=stats.form_792300_studying)
                ws.cell(row=row, column=4, value=round(stats.conversion_percentage, 2))
                ws.cell(row=row, column=5, value=prize)
                
                # Выделяем призеров
                if prize:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFD700", end_color="FFD700", fill_type="solid"
                        )
                
                row += 1
            
            # Добавляем пустую строку между группами
            row += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_branch_summary_sheet(self, wb: Workbook) -> None:
        """Создает лист со статистикой по филиалам."""
        ws = wb.create_sheet("Статистика по филиалам")
        
        # Заголовки
        headers = [
            "🏢 Филиал",
            "👴 Ст: Всего",
            "🎓 Ст: Учится",
            "📊 Ст: %",
            "👶 Нов.: Всего",
            "🎓 Нов.: Учится", 
            "📊 Нов.: %",
            "🏆 Итого %",
            "🎁 Приз"
        ]
        
        # Применяем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Сортируем филиалы по итоговому проценту (по убыванию)
        sorted_branches = sorted(
            self.branches_stats.values(),
            key=lambda x: x.total_percentage,
            reverse=True
        )
        
        # Определяем призы для топ-5 филиалов
        branch_prizes = [
            "📱 Interactive Display",      # 1 место
            "☕ Кофемашина + 2 кг кофе",   # 2 место  
            "☕ Кофемашина",               # 3 место
            "💰 20 000 руб.",             # 4 место
            "💰 10 000 руб."              # 5 место
        ]
        
        # Данные по филиалам
        row = 2
        for i, branch_stats in enumerate(sorted_branches):
            # Определяем приз
            prize = ""
            if i < len(branch_prizes):
                prize = branch_prizes[i]
            
            ws.cell(row=row, column=1, value=branch_stats.name)
            ws.cell(row=row, column=2, value=branch_stats.form_2304918_total)
            ws.cell(row=row, column=3, value=branch_stats.form_2304918_studying)
            ws.cell(row=row, column=4, value=round(branch_stats.return_percentage, 2))
            ws.cell(row=row, column=5, value=branch_stats.form_792300_total)
            ws.cell(row=row, column=6, value=branch_stats.form_792300_studying)
            ws.cell(row=row, column=7, value=round(branch_stats.conversion_percentage, 2))
            ws.cell(row=row, column=8, value=round(branch_stats.total_percentage, 2))
            ws.cell(row=row, column=9, value=prize)
            
            # Выделяем призеров ярким желтым
            if prize:
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFD700", end_color="FFD700", fill_type="solid"
                    )
            
            row += 1
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Уменьшили максимальную ширину с 50 до 30
            ws.column_dimensions[column_letter].width = adjusted_width
    
    
    async def run_analysis(self) -> None:
        """Запускает полный анализ данных."""
        # Создаем папку для отчетов если не существует
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        
        print("Начинаем анализ данных из Pyrus...")
        print(f"Время начала: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Анализируем обе формы
        await self.analyze_form_2304918()
        await self.analyze_form_792300()
        
        # Выводим краткую статистику
        print("\n=== КРАТКАЯ СТАТИСТИКА ===")
        total_teachers = len(self.teachers_stats)
        print(f"Всего преподавателей: {total_teachers}")
        
        # Статистика по группам старичков
        oldies_groups = {"6-15": 0, "16-34": 0, "35+": 0, "< 6": 0}
        for stats in self.teachers_stats.values():
            student_count = stats.form_2304918_total
            if 6 <= student_count <= 15:
                oldies_groups["6-15"] += 1
            elif 16 <= student_count <= 34:
                oldies_groups["16-34"] += 1
            elif student_count >= 35:
                oldies_groups["35+"] += 1
            else:
                oldies_groups["< 6"] += 1
        
        print("\nГруппы по старичкам:")
        for group, count in oldies_groups.items():
            print(f"  {group} студентов: {count} преподавателей")
        
        # Статистика по группам БПЗ
        trial_groups = {"16+": 0, "11-15": 0, "5-10": 0, "< 5": 0}
        for stats in self.teachers_stats.values():
            bpz_count = stats.form_792300_total
            if 5 <= bpz_count <= 10:
                trial_groups["5-10"] += 1
            elif 11 <= bpz_count <= 15:
                trial_groups["11-15"] += 1
            elif bpz_count >= 16:
                trial_groups["16+"] += 1
            else:
                trial_groups["< 5"] += 1
        
        print("\nГруппы по БПЗ студентам:")
        for group, count in trial_groups.items():
            print(f"  {group} БПЗ студентов: {count} преподавателей")
        
        # Статистика по филиалам
        if self.branches_stats:
            print(f"\nВсего филиалов: {len(self.branches_stats)}")
            print("\nТоп-5 филиалов по итоговому проценту:")
            sorted_branches = sorted(
                self.branches_stats.values(),
                key=lambda x: x.total_percentage,
                reverse=True
            )
            for i, branch_stats in enumerate(sorted_branches[:5], 1):
                print(f"{i}. {branch_stats.name}: {branch_stats.total_percentage:.2f}% "
                      f"(возврат: {branch_stats.return_percentage:.2f}%, "
                      f"конверсия: {branch_stats.conversion_percentage:.2f}%) "
                      f"[{branch_stats.form_2304918_total} старички, {branch_stats.form_792300_total} новый клиент]")
        
        if total_teachers > 0:
            print("\nТоп-5 преподавателей по итоговому проценту:")
            sorted_teachers = sorted(
                self.teachers_stats.values(),
                key=lambda x: x.total_percentage,
                reverse=True
            )
            for i, stats in enumerate(sorted_teachers[:5], 1):
                print(f"{i}. {stats.name}: {stats.total_percentage:.2f}% "
                      f"(возврат: {stats.return_percentage:.2f}%, "
                      f"конверсия: {stats.conversion_percentage:.2f}%) "
                      f"[{stats.form_2304918_total} форм 2304918]")
        
        # Создаем Excel отчет с категориями по вкладкам
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reports/pyrus_teacher_report_{timestamp}.xlsx"
        self.create_excel_reports(filename)
        
        print(f"\nАнализ завершен: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


async def main():
    """Главная функция скрипта."""
    try:
        analyzer = PyrusDataAnalyzer()
        await analyzer.run_analysis()
    except KeyboardInterrupt:
        print("\nОтмена выполнения пользователем.")
        sys.exit(1)
    except Exception as e:
        print(f"Ошибка выполнения: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    # Проверяем переменные окружения
    if not os.getenv("PYRUS_LOGIN") or not os.getenv("PYRUS_SECURITY_KEY"):
        print("ОШИБКА: Не установлены переменные окружения PYRUS_LOGIN и PYRUS_SECURITY_KEY")
        print("Убедитесь, что .env файл настроен корректно.")
        sys.exit(1)
    
    asyncio.run(main())
