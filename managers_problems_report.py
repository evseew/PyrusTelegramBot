#!/usr/bin/env python3
"""
Скрипт для создания Excel отчета по проблемным формам для руководителей.

Анализирует проблемные формы 2304918 и 792300, группирует по руководителям и филиалам:
- Галина Александровна: ЧТЗ, Копейск, Славы, Чурилово
- Валерия Владимировна: ЧМЗ, Парковый, Центр, Онлайн  
- Екатерина Евгеньевна: Чичерина, Ленинский, Макеева, Академ, Кашириных

Создает Excel файл с:
- Отдельные вкладки для каждого руководителя
- Две таблицы на каждой вкладке:
  1. Проблемы с правилами форм (классические нарушения)
  2. Контроль посещаемости (клиенты, не дошедшие до занятий)
- Список проблемных задач с указанием статуса регистрации преподавателя
- Сводная статистика по каждой группе с разбивкой по типам проблем

Логика контроля посещаемости:
- Форма 2304918: показывать если в полях 27,32,50,57 есть "Нет", убирать когда в поле 57 появится "Да"
- Форма 792300: показывать если в полях 183,198 есть "Нет", убирать когда в поле 198 появится "Да"
"""

import asyncio
import sys
import os
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, timedelta
from collections import defaultdict
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import pytz

# Загружаем переменные окружения из .env файла
load_dotenv()

# Добавляем app в путь для импорта
sys.path.append(str(Path(__file__).parent / "app"))

from app.pyrus_client import PyrusClient
from app.db import db
from app.rules.form_2304918 import check_rules as check_rules_2304918, _get_field_value
from app.rules.form_792300 import check_rules as check_rules_792300
from app.schedulers.form_checks import (
    _extract_teacher_full_name, 
    _extract_teacher_user_id,
    _build_fields_meta,
    _format_errors_brief,
    _fuzzy_find_user_by_full_name
)


# Маппинг филиалов к руководителям
MANAGERS_MAPPING = {
    "Галина Александровна": ["ЧТЗ", "Копейск", "Славы", "Чурилово"],
    "Валерия Владимировна": ["ЧМЗ", "Парковый", "Центр", "Онлайн"],
    "Екатерина Евгеньевна": ["Чичерина", "Ленинский", "Макеева", "Академ", "Кашириных"]
}

# ID полей для филиалов в формах
BRANCH_FIELD_IDS = {
    2304918: 5,  # Поле филиала в форме 2304918
    792300: 226  # Поле филиала в форме 792300
}

# Поля для контроля посещаемости
ATTENDANCE_CONTROL_FIELDS = {
    2304918: {
        "check_fields": [27, 32, 50, 57],  # Проверяем "Нет" в любом из этих полей
        "completion_field": 64,  # УЧИТСЯ (заполняет СО) - убираем из контроля когда галочка стоит
        "exclusion_field": 25,  # Статус выхода - исключаем если "Не выйдет" или "Выходит на ИЗ"
        "exclusion_values": ["не выйдет", "выходит на из"]
    },
    792300: {
        "check_fields": [183, 198],  # Пришел на 1ое/2ое занятие? - проверяем "Нет" 
        "completion_field": 187  # УЧИТСЯ - убираем из контроля когда галочка стоит
    }
}


@dataclass
class ProblemTask:
    """Информация о проблемной задаче."""
    task_id: int
    form_id: int
    form_name: str
    teacher_name: str
    teacher_registered: bool
    branch_name: str
    manager_name: str
    errors: List[str]
    brief_errors: str
    student_name: str = ""
    pyrus_link: str = ""
    problem_type: str = "rules"  # "rules" или "attendance"
    lesson_number: str = ""  # "1-е занятие", "2-е занятие" для attendance


class ManagersProblemsAnalyzer:
    """Анализатор проблемных форм для руководителей."""
    
    def __init__(self):
        self.client = PyrusClient()
        self.registered_teachers = set()  # Множество зарегистрированных учителей (user_id)
        self.teacher_names_map = {}  # user_id -> full_name для зарегистрированных
        self.problems_by_manager: Dict[str, List[ProblemTask]] = defaultdict(list)
        self.attendance_by_manager: Dict[str, List[ProblemTask]] = defaultdict(list)  # Отдельно для "не дошли до урока"
        
    def _get_target_day(self) -> str:
        """Получить целевую дату для проверки (вчера)."""
        tz = pytz.timezone(os.getenv("TZ", "Asia/Yekaterinburg"))
        now_local = datetime.now(tz)
        yesterday = (now_local - timedelta(days=1)).date()
        return yesterday.isoformat()
    
    def _load_registered_teachers(self) -> None:
        """Загружает список зарегистрированных преподавателей из БД."""
        try:
            users = db.get_all_users()
            for user in users:
                self.registered_teachers.add(user.user_id)
                self.teacher_names_map[user.user_id] = user.full_name or f"User {user.user_id}"
            print(f"Загружено {len(self.registered_teachers)} зарегистрированных преподавателей")
        except Exception as e:
            print(f"⚠️ Ошибка загрузки пользователей из БД: {e}")
            print("⚠️ Продолжаем работу без данных о регистрации преподавателей")
            # Очищаем списки для корректной работы без БД
            self.registered_teachers.clear()
            self.teacher_names_map.clear()
    
    def _extract_branch_name(self, task_fields: List[Dict[str, Any]], field_id: int) -> str:
        """Извлекает название филиала из поля справочника."""
        value = _get_field_value(task_fields, field_id)
        
        if isinstance(value, dict):
            # Проверяем массив values - основной способ для справочника филиалов
            values = value.get("values")
            if isinstance(values, list) and len(values) > 0:
                branch_name = values[0]  # Берем первое значение
                if isinstance(branch_name, str) and branch_name.strip():
                    return branch_name.strip()
            
            # Проверяем rows если values не найден  
            rows = value.get("rows")
            if isinstance(rows, list) and len(rows) > 0 and isinstance(rows[0], list) and len(rows[0]) > 0:
                branch_name = rows[0][0]  # Берем первую ячейку первой строки
                if isinstance(branch_name, str) and branch_name.strip():
                    return branch_name.strip()
            
            # Для обычных справочников обычно есть поле text или name
            for key in ("text", "name", "value"):
                branch_val = value.get(key)
                if isinstance(branch_val, str) and branch_val.strip():
                    return branch_val.strip()
        
        if isinstance(value, str):
            return value.strip()
        
        return "Неизвестный филиал"
    
    def _find_manager_by_branch(self, branch_name: str) -> Optional[str]:
        """Определяет руководителя по названию филиала."""
        for manager, branches in MANAGERS_MAPPING.items():
            for branch_pattern in branches:
                if branch_pattern.lower() in branch_name.lower():
                    return manager
        return None
    
    def _extract_student_name(self, task_fields: List[Dict[str, Any]], form_id: int) -> str:
        """Извлекает ФИО студента из соответствующего поля формы."""
        if form_id == 2304918:
            student_field_id = 2  # Поле ФИО студента в форме 2304918
        elif form_id == 792300:
            student_field_id = 73  # Поле ФИО студента в форме 792300
        else:
            return "Неизвестный студент"
        
        value = _get_field_value(task_fields, student_field_id)
        
        if isinstance(value, dict):
            # Поддержка person-объекта: first_name/last_name
            first_name = value.get("first_name", "")
            last_name = value.get("last_name", "")
            if isinstance(first_name, str) or isinstance(last_name, str):
                full_name = f"{(first_name or '').strip()} {(last_name or '').strip()}".strip()
                if full_name:
                    return full_name
            
            # Для обычных текстовых полей
            for key in ("text", "name", "value"):
                name_val = value.get(key)
                if isinstance(name_val, str) and name_val.strip():
                    return name_val.strip()
        
        if isinstance(value, str):
            return value.strip()
        
        return "Неизвестный студент"
    
    def _is_teacher_registered(self, task_fields: List[Dict[str, Any]], form_id: int) -> Tuple[str, bool]:
        """
        Проверяет статус регистрации преподавателя.
        
        Returns:
            Tuple[str, bool]: (teacher_name, is_registered)
        """
        if form_id == 2304918:
            teacher_field_id = 8  # Основное поле преподавателя
            teacher_rule3_field_id = 49  # Поле для правила 3
            
            # Проверяем, заполнено ли поле 49 (новый преподаватель)
            rule3_value = _get_field_value(task_fields, teacher_rule3_field_id)
            use_rule3_field = rule3_value is not None and str(rule3_value).strip()
            
            if use_rule3_field:
                # Используем поле 49
                teacher_user_id = _extract_teacher_user_id(task_fields, teacher_rule3_field_id)
                teacher_name = _extract_teacher_full_name(task_fields, teacher_rule3_field_id)
            else:
                # Используем основное поле 8
                teacher_user_id = _extract_teacher_user_id(task_fields, teacher_field_id)
                teacher_name = _extract_teacher_full_name(task_fields, teacher_field_id)
        else:
            # Для формы 792300
            from app.rules.form_792300 import _extract_teacher_full_name as f792300_extract_teacher_full_name
            from app.rules.form_792300 import _extract_teacher_user_id as f792300_extract_teacher_user_id
            teacher_field_id = 142
            teacher_user_id = f792300_extract_teacher_user_id(task_fields, teacher_field_id)
            teacher_name = f792300_extract_teacher_full_name(task_fields, teacher_field_id)
        
        # Проверяем регистрацию (только если есть данные о зарегистрированных пользователях)
        if isinstance(teacher_user_id, int) and self.registered_teachers and teacher_user_id in self.registered_teachers:
            return teacher_name or "Неизвестный преподаватель", True
        
        # Пробуем fuzzy-поиск для форм, которые это поддерживают (только если есть данные о пользователях)
        if form_id == 2304918 and teacher_name and self.registered_teachers:
            _, found_user_id = _fuzzy_find_user_by_full_name(teacher_name, threshold=0.85)
            if found_user_id and found_user_id in self.registered_teachers:
                return teacher_name, True
        
        # Если нет данных о регистрации или преподаватель не найден
        return teacher_name or "Неизвестный преподаватель", False
    
    def _check_attendance_problems(self, task_fields: List[Dict[str, Any]], form_id: int) -> tuple[bool, str]:
        """
        Проверяет, требуется ли контроль посещаемости для данной задачи.
        
        Returns:
            tuple[bool, str]: (нужен_ли_контроль, детали_занятия)
        """
        control_config = ATTENDANCE_CONTROL_FIELDS.get(form_id)
        if not control_config:
            return False, ""
        
        # Проверяем поле завершения - если там "Да", то контроль не нужен
        completion_field_id = control_config["completion_field"]
        completion_value = _get_field_value(task_fields, completion_field_id)
        
        
        # Проверяем галочку "УЧИТСЯ" - если стоит, то контроль не нужен
        if isinstance(completion_value, dict):
            # Проверяем поле типа checkbox
            checkmark = completion_value.get("checkmark")
            if checkmark == "checked":
                return False, ""
            # Дополнительная проверка на текстовые значения
            text_value = completion_value.get("text", "").lower()
            if text_value in ["да", "yes"]:
                return False, ""
        elif isinstance(completion_value, str):
            if completion_value.lower() in ["да", "yes", "checked"]:
                return False, ""
        
        # Проверяем поле исключения (только для формы 2304918)
        exclusion_field_id = control_config.get("exclusion_field")
        if exclusion_field_id:
            exclusion_value = _get_field_value(task_fields, exclusion_field_id)
            exclusion_values = control_config.get("exclusion_values", [])
            
            if exclusion_value:
                # Проверяем различные типы значений
                if isinstance(exclusion_value, dict):
                    # Проверяем choice_names для выпадающих списков
                    choice_names = exclusion_value.get("choice_names", [])
                    if isinstance(choice_names, list):
                        for choice in choice_names:
                            if isinstance(choice, str):
                                if any(ev.lower() in choice.lower() for ev in exclusion_values):
                                    return False, ""
                    
                    # Проверяем text для текстовых полей
                    text_value = exclusion_value.get("text", "").lower()
                    if any(ev.lower() in text_value for ev in exclusion_values):
                        return False, ""
                elif isinstance(exclusion_value, str):
                    if any(ev.lower() in exclusion_value.lower() for ev in exclusion_values):
                        return False, ""
                elif isinstance(exclusion_value, list):
                    for item in exclusion_value:
                        if isinstance(item, str):
                            if any(ev.lower() in item.lower() for ev in exclusion_values):
                                return False, ""
        
        # Проверяем наличие "Нет" в контрольных полях
        check_fields = control_config["check_fields"]
        
        # Определяем соответствие полей занятиям
        lesson_mapping = {}
        if form_id == 2304918:
            lesson_mapping = {
                27: "1-е занятие",    # ATTENDED_1_ID
                32: "1-е занятие",    # ATTENDED_OTHER_GROUP_ID (тоже про первое)
                50: "1-е занятие",    # ATTENDED_FIRST_ID
                57: "2-е занятие"     # ATTENDED_SECOND_ID
            }
        elif form_id == 792300:
            lesson_mapping = {
                183: "1-е занятие",   # Пришел на 1ое занятие?
                198: "2-е занятие"    # Пришел на 2ое занятие?
            }
        
        for field_id in check_fields:
            field_value = _get_field_value(task_fields, field_id)
            
            # Проверяем choice_names - все поля с "Нет" имеют такую структуру
            if isinstance(field_value, dict):
                choice_names = field_value.get("choice_names", [])
                if isinstance(choice_names, list):
                    for choice in choice_names:
                        if isinstance(choice, str) and choice.lower() in ["нет", "no"]:
                            lesson_detail = lesson_mapping.get(field_id, "занятие")
                            return True, f"не дошёл до {lesson_detail}"
        
        return False, ""
    
    async def analyze_form(self, form_id: int) -> None:
        """Анализирует проблемные задачи для указанной формы."""
        print(f"Анализ формы {form_id}...")
        
        # Получаем метаданные формы
        form_meta = await self.client.get_form_meta(form_id)
        if not form_meta:
            print(f"⚠️ Не удалось получить метаданные формы {form_id}")
            return
        
        fields_meta = _build_fields_meta(form_meta)
        form_name = form_meta.get("name") or f"Форма {form_id}"
        target_day = self._get_target_day()
        
        # Выбираем функцию проверки правил
        check_rules = check_rules_2304918 if form_id == 2304918 else check_rules_792300
        branch_field_id = BRANCH_FIELD_IDS.get(form_id)
        
        if not branch_field_id:
            print(f"⚠️ Не найден ID поля филиала для формы {form_id}")
            return
        
        task_count = 0
        problem_count = 0
        
        async for task in self.client.iter_register_tasks(form_id, include_archived=False):
            task_count += 1
            if task_count % 100 == 0:
                print(f"Обработано {task_count} задач формы {form_id}...")
            
            
            
            
            
            
            task_id = task.get("id") or task.get("task_id")
            task_fields = task.get("fields", [])
            
            # Получаем заголовок задачи
            task_title = (task.get("subject") or task.get("text") or "").strip()
            if not task_title:
                # Fallback: поле id=1 в fields
                for f in task_fields or []:
                    if f.get("id") == 1:
                        val = f.get("value") or {}
                        if isinstance(val, dict):
                            task_title = str(val.get("text") or val.get("value") or val.get("name") or f.get("name") or "Задача").strip()
                        elif isinstance(val, str):
                            task_title = val.strip() or (f.get("name") or "Задача").strip()
                        else:
                            task_title = (f.get("name") or "Задача").strip()
                        break
            if not task_title:
                task_title = f"Задача #{task_id}"
            
            # Извлекаем данные о преподавателе, студенте и филиале (общие для обоих типов проблем)
            teacher_name, is_registered = self._is_teacher_registered(task_fields, form_id)
            student_name = self._extract_student_name(task_fields, form_id)
            branch_name = self._extract_branch_name(task_fields, branch_field_id)
            manager_name = self._find_manager_by_branch(branch_name)
            
            # Пропускаем задачи, которые не относятся к отслеживаемым филиалам
            if not manager_name:
                continue
            
            # Проверяем правила (проблемы типа "rules")
            errors_map = check_rules(fields_meta, task_fields, target_day, "yesterday12")
            general_errors = errors_map.get("general", [])
            rule3_errors = errors_map.get("rule3", [])
            all_errors = general_errors + rule3_errors
            
            if all_errors:
                problem_count += 1
                
                # Создаем объект проблемной задачи типа "rules"
                problem_task = ProblemTask(
                    task_id=task_id,
                    form_id=form_id,
                    form_name=form_name,
                    teacher_name=teacher_name,
                    teacher_registered=is_registered,
                    branch_name=branch_name,
                    manager_name=manager_name,
                    errors=all_errors,
                    brief_errors=_format_errors_brief(all_errors),
                    student_name=student_name,
                    pyrus_link=f"https://pyrus.com/t#id{task_id}",
                    problem_type="rules"
                )
                
                self.problems_by_manager[manager_name].append(problem_task)
            
            # Проверяем проблемы с посещаемостью (проблемы типа "attendance")
            needs_attendance_control, lesson_detail = self._check_attendance_problems(task_fields, form_id)
            if needs_attendance_control:
                problem_count += 1
                
                # Создаем объект проблемной задачи типа "attendance"
                attendance_problem = ProblemTask(
                    task_id=task_id,
                    form_id=form_id,
                    form_name=form_name,
                    teacher_name=teacher_name,
                    teacher_registered=is_registered,
                    branch_name=branch_name,
                    manager_name=manager_name,
                    errors=[lesson_detail],
                    brief_errors=lesson_detail,
                    student_name=student_name,
                    pyrus_link=f"https://pyrus.com/t#id{task_id}",
                    problem_type="attendance",
                    lesson_number=lesson_detail
                )
                
                self.attendance_by_manager[manager_name].append(attendance_problem)
        
        print(f"Завершен анализ формы {form_id}. Обработано {task_count} задач, найдено {problem_count} проблемных.")
    
    def create_excel_report(self, filename: str = "managers_problems_report.xlsx") -> None:
        """Создает Excel отчет с проблемными задачами по руководителям."""
        print(f"Создание Excel отчета: {filename}")
        
        wb = Workbook()
        # Удаляем дефолтный лист
        wb.remove(wb.active)
        
        # Создаем вкладки для каждого руководителя
        for manager_name in MANAGERS_MAPPING.keys():
            rules_problems = self.problems_by_manager[manager_name]
            attendance_problems = self.attendance_by_manager[manager_name]
            
            # Вкладка с проблемами правил
            if rules_problems:
                print(f"Создание вкладки '{manager_name}': {len(rules_problems)} проблемных задач")
                self._create_manager_sheet(wb, manager_name, rules_problems)
            
            # Отдельная вкладка "не дошли до урока"
            if attendance_problems:
                attendance_sheet_name = f"{manager_name}_не_дошли"
                print(f"Создание вкладки '{attendance_sheet_name}': {len(attendance_problems)} задач")
                self._create_attendance_sheet(wb, attendance_sheet_name, attendance_problems, manager_name)
        
        # Создаем сводную вкладку
        if any(self.problems_by_manager.values()) or any(self.attendance_by_manager.values()):
            print("Создание сводной вкладки")
            self._create_summary_sheet(wb)
        
        # Сохраняем файл
        wb.save(filename)
        print(f"✅ Отчет сохранен: {filename}")
    
    def _create_manager_sheet(self, wb: Workbook, manager_name: str, problems: List[ProblemTask]) -> None:
        """Создает лист с проблемными задачами для конкретного руководителя."""
        # Укорачиваем имя для названия вкладки (Excel ограничение в 31 символ)
        sheet_name = manager_name.replace(" ", "_")[:31]
        ws = wb.create_sheet(sheet_name)
        
        # Разделяем проблемы по типам
        rules_problems = [p for p in problems if p.problem_type == "rules"]
        attendance_problems = [p for p in problems if p.problem_type == "attendance"]
        
        row = 1
        
        # Создаем первую таблицу - проблемы с правилами
        if rules_problems:
            row = self._create_problems_table(ws, rules_problems, "ПРОБЛЕМЫ С ФОРМАМИ", row, "FF6B6B")
            row += 2  # Пустые строки между таблицами
        
        # Создаем вторую таблицу - не дошли до урока
        if attendance_problems:
            row = self._create_problems_table(ws, attendance_problems, "НЕ ДОШЛИ ДО УРОКА", row, "4ECDC4")
            row += 2  # Пустые строки перед статистикой
        
        # Статистика внизу
        if rules_problems or attendance_problems:
            self._create_statistics_section(ws, problems, row)
        
        # Автоширина колонок
        self._adjust_column_widths(ws)
    
    def _create_attendance_sheet(self, wb: Workbook, sheet_name: str, attendance_problems: List[ProblemTask], manager_name: str) -> None:
        """Создает отдельную вкладку 'не дошли до урока' для руководителя."""
        # Укорачиваем имя для названия вкладки (Excel ограничение в 31 символ)
        safe_sheet_name = sheet_name.replace(" ", "_")[:31]
        ws = wb.create_sheet(safe_sheet_name)
        
        row = 1
        
        # Заголовок вкладки
        ws.cell(row=row, column=1, value=f"НЕ ДОШЛИ ДО УРОКА - {manager_name}")
        merge_range = f"A{row}:E{row}"  # 5 колонок для attendance
        ws.merge_cells(merge_range)
        title_cell = ws.cell(row=row, column=1)
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2
        
        # Создаем таблицу с проблемными задачами
        row = self._create_problems_table(ws, attendance_problems, "СТУДЕНТЫ, НЕ ДОШЕДШИЕ ДО УРОКА", row, "FFB84D")
        row += 2
        
        # Статистика
        unique_teachers = len({p.teacher_name for p in attendance_problems})
        registered = len({p.teacher_name for p in attendance_problems if p.teacher_registered})
        unregistered = len({p.teacher_name for p in attendance_problems if not p.teacher_registered})
        branches = len({p.branch_name for p in attendance_problems})
        
        stats = [
            f"Всего студентов не дошедших: {len(attendance_problems)}",
            f"Уникальных преподавателей: {unique_teachers}",
            f"Зарегистрированных преподавателей: {registered}",
            f"Незарегистрированных преподавателей: {unregistered}",
            f"Филиалов: {branches}",
            "",
            "ВНИМАНИЕ: Задачи остаются в этом списке до постановки галочки 'УЧИТСЯ':"
        ]
        
        # Добавляем информацию о полях завершения для разных форм
        forms_info = []
        forms_in_list = {p.form_id for p in attendance_problems}
        if 2304918 in forms_in_list:
            forms_info.extend([
                "- Форма 2304918: поле 64 (УЧИТСЯ) ИЛИ",
                "- Форма 2304918: поле 25 (статус 'Не выйдет'/'Выходит на ИЗ')"
            ])
        if 792300 in forms_in_list:
            forms_info.append("- Форма 792300: поле 187 (УЧИТСЯ)")
        
        stats.extend(forms_info)
        
        for stat in stats:
            ws.cell(row=row, column=1, value=stat)
            if stat.startswith("ВНИМАНИЕ") or stat.startswith("-"):
                ws.cell(row=row, column=1).font = Font(bold=True, color="CC0000")
            row += 1
        
        # Автоширина колонок
        self._adjust_column_widths(ws)
    
    def _create_problems_table(self, ws, problems: List[ProblemTask], table_title: str, start_row: int, header_color: str) -> int:
        """Создает таблицу с проблемными задачами определенного типа."""
        row = start_row
        
        # Заголовок таблицы (разная ширина для разных типов)
        ws.cell(row=row, column=1, value=table_title)
        if problems[0].problem_type == "attendance":
            merge_range = f"A{row}:E{row}"  # 5 колонок для attendance
        else:
            merge_range = f"A{row}:F{row}"  # 6 колонок для rules
        ws.merge_cells(merge_range)
        title_cell = ws.cell(row=row, column=1)
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Заголовки столбцов (разные для rules и attendance)
        if problems[0].problem_type == "attendance":
            headers = [
                "Форма",
                "ФИО студента", 
                "Преподаватель",
                "Проблема",
                "Ссылка"
            ]
        else:
            headers = [
                "Форма",
                "ФИО студента", 
                "Преподаватель",
                "Статус регистрации",
                "Проблемы",
                "Ссылка"
            ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        row += 1
        
        # Группируем задачи по филиалам и преподавателям
        problems_by_branch_teacher = defaultdict(lambda: defaultdict(list))
        for problem in problems:
            problems_by_branch_teacher[problem.branch_name][problem.teacher_name].append(problem)
        
        # Заполняем данные
        for branch_name in sorted(problems_by_branch_teacher.keys()):
            teachers_problems = problems_by_branch_teacher[branch_name]
            
            # Заголовок филиала (разная ширина для разных типов)
            ws.cell(row=row, column=1, value=f"Филиал: {branch_name}")
            if problems[0].problem_type == "attendance":
                merge_range = f"A{row}:E{row}"  # 5 колонок для attendance
            else:
                merge_range = f"A{row}:F{row}"  # 6 колонок для rules
            ws.merge_cells(merge_range)
            branch_cell = ws.cell(row=row, column=1)
            branch_cell.font = Font(bold=True, color="FFFFFF")
            branch_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            branch_cell.alignment = Alignment(horizontal="center")
            row += 1
            
            # Проходим по преподавателям в филиале
            for teacher_name in sorted(teachers_problems.keys()):
                teacher_problems = teachers_problems[teacher_name]
                
                # Задачи преподавателя
                for problem in sorted(teacher_problems, key=lambda x: x.task_id):
                    ws.cell(row=row, column=1, value=problem.form_name)
                    ws.cell(row=row, column=2, value=problem.student_name)
                    ws.cell(row=row, column=3, value=problem.teacher_name)
                    
                    if problem.problem_type == "attendance":
                        # Для attendance: убираем столбец регистрации
                        ws.cell(row=row, column=4, value=problem.brief_errors)
                        ws.cell(row=row, column=5, value=problem.pyrus_link)
                        col_count = 5
                    else:
                        # Для rules: оставляем столбец регистрации
                        status_text = "✅ Зарегистрирован" if problem.teacher_registered else "❌ Не зарегистрирован"
                        status_color = "D5E8D4" if problem.teacher_registered else "F8CECC"
                        
                        status_cell = ws.cell(row=row, column=4, value=status_text)
                        status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                        
                        ws.cell(row=row, column=5, value=problem.brief_errors)
                        ws.cell(row=row, column=6, value=problem.pyrus_link)
                        col_count = 6
                    
                    # Стилизация строки
                    for col in range(1, col_count + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        # Ссылка всегда в последней колонке
                        if col == col_count:
                            cell.font = Font(color="0563C1", underline="single")
                    
                    row += 1
            
            # Пустая строка между филиалами
            row += 1
        
        return row
    
    def _create_statistics_section(self, ws, problems: List[ProblemTask], start_row: int) -> int:
        """Создает секцию со статистикой."""
        row = start_row
        
        # Заголовок статистики
        ws.cell(row=row, column=1, value="СТАТИСТИКА:")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        # Разделяем статистику по типам проблем
        rules_problems = [p for p in problems if p.problem_type == "rules"]
        attendance_problems = [p for p in problems if p.problem_type == "attendance"]
        
        # Общая статистика
        total_tasks = len(problems)
        unique_teachers = len({p.teacher_name for p in problems})
        registered_teachers = len({p.teacher_name for p in problems if p.teacher_registered})
        unregistered_teachers = len({p.teacher_name for p in problems if not p.teacher_registered})
        
        # Статистика по проблемам с правилами
        rules_tasks = len(rules_problems)
        rules_teachers = len({p.teacher_name for p in rules_problems})
        
        # Статистика по "не дошли до урока"
        attendance_tasks = len(attendance_problems)
        attendance_teachers = len({p.teacher_name for p in attendance_problems})
        
        # Филиалы
        branches = {p.branch_name for p in problems}
        
        stats = [
            "=== ОБЩАЯ СТАТИСТИКА ===",
            f"Всего задач: {total_tasks}",
            f"Уникальных преподавателей: {unique_teachers}",
            f"Зарегистрированных преподавателей: {registered_teachers}",
            f"Незарегистрированных преподавателей: {unregistered_teachers}",
            f"Филиалов: {len(branches)}",
            "",
            "=== ПО ТИПАМ ПРОБЛЕМ ===",
            f"Проблемы с формами: {rules_tasks} задач, {rules_teachers} преподавателей",
            f"Не дошли до урока: {attendance_tasks} задач, {attendance_teachers} преподавателей"
        ]
        
        for stat in stats:
            ws.cell(row=row, column=1, value=stat)
            if stat.startswith("==="):
                ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        return row
    
    def _adjust_column_widths(self, ws) -> None:
        """Автоматически подгоняет ширину колонок."""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Пропускаем объединенные ячейки
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Создает сводную вкладку со статистикой по всем руководителям."""
        ws = wb.create_sheet("Сводка")
        
        # Заголовок
        ws.cell(row=1, column=1, value="СВОДНАЯ СТАТИСТИКА ПО РУКОВОДИТЕЛЯМ")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells("A1:H1")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        
        # Заголовки таблицы
        headers = [
            "Руководитель",
            "Всего задач",
            "Проблемы с формами",
            "Не дошли до урока",
            "Уникальных преподавателей",
            "Зарегистрированных",
            "Незарегистрированных",
            "Филиалов"
        ]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row += 1
        
        # Статистика по руководителям
        total_tasks_all = 0
        total_rules_all = 0
        total_attendance_all = 0
        total_teachers_all = set()
        total_registered_all = set()
        total_unregistered_all = set()
        total_branches_all = set()
        
        for manager_name in MANAGERS_MAPPING.keys():
            rules_problems = self.problems_by_manager[manager_name]
            attendance_problems = self.attendance_by_manager[manager_name]
            
            if not rules_problems and not attendance_problems:
                continue
            
            # Объединяем все проблемы для общей статистики
            all_problems = rules_problems + attendance_problems
            
            total_tasks = len(all_problems)
            rules_count = len(rules_problems)
            attendance_count = len(attendance_problems)
            unique_teachers = {p.teacher_name for p in all_problems}
            registered_teachers = {p.teacher_name for p in all_problems if p.teacher_registered}
            unregistered_teachers = {p.teacher_name for p in all_problems if not p.teacher_registered}
            branches = {p.branch_name for p in all_problems}
            
            ws.cell(row=row, column=1, value=manager_name)
            ws.cell(row=row, column=2, value=total_tasks)
            ws.cell(row=row, column=3, value=rules_count)
            ws.cell(row=row, column=4, value=attendance_count)
            ws.cell(row=row, column=5, value=len(unique_teachers))
            ws.cell(row=row, column=6, value=len(registered_teachers))
            ws.cell(row=row, column=7, value=len(unregistered_teachers))
            ws.cell(row=row, column=8, value=len(branches))
            
            # Обновляем общую статистику
            total_tasks_all += total_tasks
            total_rules_all += rules_count
            total_attendance_all += attendance_count
            total_teachers_all.update(unique_teachers)
            total_registered_all.update(registered_teachers)
            total_unregistered_all.update(unregistered_teachers)
            total_branches_all.update(branches)
            
            row += 1
        
        # Итоговая строка
        row += 1
        ws.cell(row=row, column=1, value="ИТОГО:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_tasks_all)
        ws.cell(row=row, column=3, value=total_rules_all)
        ws.cell(row=row, column=4, value=total_attendance_all)
        ws.cell(row=row, column=5, value=len(total_teachers_all))
        ws.cell(row=row, column=6, value=len(total_registered_all))
        ws.cell(row=row, column=7, value=len(total_unregistered_all))
        ws.cell(row=row, column=8, value=len(total_branches_all))
        
        # Стилизация итоговой строки
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Пропускаем объединенные ячейки
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    async def run_analysis(self) -> None:
        """Запускает полный анализ проблемных форм."""
        # Создаем папку для отчетов если не существует
        reports_dir = Path("reports")
        reports_dir.mkdir(exist_ok=True)
        
        print("Начинаем анализ проблемных форм для руководителей...")
        print(f"Время начала: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Загружаем зарегистрированных преподавателей
        self._load_registered_teachers()
        
        # Анализируем обе формы
        await self.analyze_form(2304918)
        await self.analyze_form(792300)
        
        # Выводим краткую статистику
        print("\n=== КРАТКАЯ СТАТИСТИКА ===")
        for manager_name in MANAGERS_MAPPING.keys():
            rules_problems = self.problems_by_manager[manager_name]
            attendance_problems = self.attendance_by_manager[manager_name]
            
            if rules_problems or attendance_problems:
                all_problems = rules_problems + attendance_problems
                unique_teachers = len({p.teacher_name for p in all_problems})
                registered = len({p.teacher_name for p in all_problems if p.teacher_registered})
                unregistered = len({p.teacher_name for p in all_problems if not p.teacher_registered})
                print(f"{manager_name}: {len(all_problems)} задач ({len(rules_problems)} правила, "
                      f"{len(attendance_problems)} не дошли), {unique_teachers} преподавателей "
                      f"(зарег: {registered}, не зарег: {unregistered})")
        
        # Создаем Excel отчет
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reports/managers_problems_report_{timestamp}.xlsx"
        self.create_excel_report(filename)
        
        print(f"\nАнализ завершен: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


async def main():
    """Главная функция скрипта."""
    try:
        analyzer = ManagersProblemsAnalyzer()
        await analyzer.run_analysis()
    except KeyboardInterrupt:
        print("\nОтмена выполнения пользователем.")
        sys.exit(1)
    except Exception as e:
        print(f"Ошибка выполнения: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    # Проверяем переменные окружения
    if not os.getenv("PYRUS_LOGIN") or not os.getenv("PYRUS_SECURITY_KEY"):
        print("ОШИБКА: Не установлены переменные окружения PYRUS_LOGIN и PYRUS_SECURITY_KEY")
        print("Убедитесь, что .env файл настроен корректно.")
        sys.exit(1)
    
    asyncio.run(main())
