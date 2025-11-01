#!/usr/bin/env python3
"""
Отчет по проверкам уроков преподавателей (форма 697550).

Формирует Excel-отчет с тремя вкладками:
1. Рейтинг филиалов по средней оценке (с призами топ-3)
2. Детали по филиалам (все преподаватели с их оценками)
3. Рейтинги по категориям, программам и курсам

Форма 697550 "Анализ урока":
  * Поле 47: Филиал (catalog)
  * Поле 48: Преподаватель (person)
  * Поле 34: Оценка за урок (multiple_choice: выпадающий список 1-10)
            ВЛОЖЕННОЕ в поле 29 "Анализ урока. Заполняет проверяющий"
  * Поле 45: Категория (multiple_choice)
  * Поле 37: Программа (catalog)
  * Поле 52: Курс (multiple_choice)
  
  ВАЖНО: Поле 34 находится во вложенной секции, метод _get_field_value
  выполняет рекурсивный поиск для извлечения значений из вложенных полей.
  Метод _extract_lesson_score поддерживает извлечение числовых значений
  из choice_names для полей типа multiple_choice.
"""

import asyncio
import sys
import os
import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# Загружаем переменные окружения из .env файла
load_dotenv()

# Добавляем app в путь для импорта
sys.path.append(str(Path(__file__).parent / "app"))

from pyrus_client import PyrusClient


class LessonCheckStats:
    """Статистика проверок уроков для преподавателя."""
    
    def __init__(self, name: str):
        self.name = name
        self.total_checks = 0
        self.total_score = 0.0  # Сумма оценок
        
        # Детальная статистика по измерениям
        self.checks_by_branch: Dict[str, List[float]] = defaultdict(list)
        self.checks_by_category: Dict[str, List[float]] = defaultdict(list)
        self.checks_by_program: Dict[str, List[float]] = defaultdict(list)
        self.checks_by_course: Dict[str, List[float]] = defaultdict(list)
        
        # Сырые данные для детализации
        self.raw_checks: List[Dict[str, Any]] = []
    
    @property
    def average_score(self) -> float:
        """Средняя оценка преподавателя."""
        if self.total_checks == 0:
            return 0.0
        return self.total_score / self.total_checks
    
    def add_check(self, score: float, branch: str, category: str, program: str, course: str, task_id: int) -> None:
        """Добавляет проверку в статистику."""
        self.total_checks += 1
        self.total_score += score
        
        # Группируем по измерениям
        self.checks_by_branch[branch].append(score)
        self.checks_by_category[category].append(score)
        self.checks_by_program[program].append(score)
        self.checks_by_course[course].append(score)
        
        # Сохраняем сырые данные
        self.raw_checks.append({
            "task_id": task_id,
            "score": score,
            "branch": branch,
            "category": category,
            "program": program,
            "course": course
        })


class BranchLessonStats:
    """Статистика проверок уроков для филиала."""
    
    def __init__(self, name: str):
        self.name = name
        self.total_checks = 0
        self.total_score = 0.0
        self.teachers: Dict[str, List[float]] = defaultdict(list)  # {teacher: [scores]}
    
    @property
    def average_score(self) -> float:
        """Средняя оценка филиала."""
        if self.total_checks == 0:
            return 0.0
        return self.total_score / self.total_checks
    
    def add_check(self, score: float, teacher_name: str) -> None:
        """Добавляет проверку в статистику филиала."""
        self.total_checks += 1
        self.total_score += score
        self.teachers[teacher_name].append(score)


class LessonChecksAnalyzer:
    """Анализатор данных проверок уроков из Pyrus."""
    
    def __init__(self, config_path: str = "lesson_checks_config.json"):
        self.client = PyrusClient()
        self.teachers_stats: Dict[str, LessonCheckStats] = {}
        self.branches_stats: Dict[str, BranchLessonStats] = {}
        
        # Загружаем конфигурацию
        self.config = self._load_config(config_path)
        
        # Счетчики для отладки
        self.total_tasks_processed = 0
        self.valid_tasks_count = 0
        self.invalid_score_count = 0
        self.excluded_by_period_count = 0
        self.excluded_teachers_count = 0
        self.excluded_branches_count = 0
        self.unknown_branch_count = 0  # Счетчик "Неизвестный филиал"
        
        # Трекинг дат для статистики
        self.earliest_date = None
        self.latest_date = None
    
    def _load_config(self, config_path: str) -> Dict[str, Any]:
        """Загружает конфигурацию из JSON файла."""
        config_file = Path(__file__).parent / config_path
        
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                print(f"✅ Конфигурация загружена из {config_path}")
                return config
        except FileNotFoundError:
            print(f"⚠️ Файл конфигурации {config_path} не найден. Используются настройки по умолчанию.")
            return self._get_default_config()
        except json.JSONDecodeError as e:
            print(f"⚠️ Ошибка парсинга {config_path}: {e}. Используются настройки по умолчанию.")
            return self._get_default_config()
    
    def _get_default_config(self) -> Dict[str, Any]:
        """Возвращает конфигурацию по умолчанию."""
        return {
            "form_id": 697550,
            "period": {"enabled": False},
            "excluded_teachers": [],
            "excluded_branches": [],
            "min_score": 1,
            "max_score": 10,
            "include_archived": True,
            "prizes": {
                "branch_top3": ["💰 15 000 руб.", "💰 10 000 руб.", "💰 5 000 руб."]
            },
            "output": {
                "directory": "reports",
                "filename_prefix": "lesson_checks_report",
                "add_timestamp": True
            }
        }
    
    def _get_field_value(self, field_list: List[Dict[str, Any]], field_id: int) -> Optional[Any]:
        """Ищет значение поля по id, рекурсивно обходя вложенные секции."""
        for f in field_list or []:
            if f.get("id") == field_id:
                return f.get("value")
            val = f.get("value")
            if isinstance(val, dict) and isinstance(val.get("fields"), list):
                nested = self._get_field_value(val.get("fields") or [], field_id)
                if nested is not None:
                    return nested
        return None
    
    def _extract_teacher_name(self, task_fields: List[Dict[str, Any]], field_id: int) -> str:
        """Извлекает ФИО преподавателя из поля справочника."""
        value = self._get_field_value(task_fields, field_id)
        
        if isinstance(value, dict):
            # Поддержка person-объекта: first_name/last_name
            first_name = value.get("first_name", "")
            last_name = value.get("last_name", "")
            if isinstance(first_name, str) or isinstance(last_name, str):
                full_name = f"{(first_name or '').strip()} {(last_name or '').strip()}".strip()
                if full_name:
                    return full_name
            
            # Для справочника сотрудников обычно есть поле text или name
            for key in ("text", "name", "value"):
                name_val = value.get(key)
                if isinstance(name_val, str) and name_val.strip():
                    return name_val.strip()
        
        if isinstance(value, str):
            return value.strip()
        
        return "Неизвестный преподаватель"
    
    def _normalize_branch_name(self, branch_name: str) -> str:
        """Нормализует название филиала для объединения данных."""
        branch_name = branch_name.lower().strip()
        
        # Только филиал Коммунистический 22 считается как Копейск
        if "коммунистический" in branch_name and "22" in branch_name:
            return "Копейск"
        
        # Возвращаем оригинальное название с заглавной буквы
        return branch_name.title()
    
    def _extract_branch_name(self, task_fields: List[Dict[str, Any]], field_id: int, debug: bool = False) -> str:
        """Извлекает название филиала из поля справочника."""
        value = self._get_field_value(task_fields, field_id)
        
        if debug and value:
            print(f"🔍 DEBUG: Поле {field_id} содержит: {json.dumps(value, ensure_ascii=False, indent=2)[:500]}")
        
        if isinstance(value, dict):
            # Проверяем массив values - основной способ для справочника филиалов
            values = value.get("values")
            if isinstance(values, list) and len(values) > 0:
                branch_name = values[0]  # Берем первое значение
                if isinstance(branch_name, str) and branch_name.strip():
                    return self._normalize_branch_name(branch_name.strip())
            
            # Проверяем rows если values не найден  
            rows = value.get("rows")
            if isinstance(rows, list) and len(rows) > 0 and isinstance(rows[0], list) and len(rows[0]) > 0:
                branch_name = rows[0][0]  # Берем первую ячейку первой строки
                if isinstance(branch_name, str) and branch_name.strip():
                    return self._normalize_branch_name(branch_name.strip())
            
            # Для обычных справочников обычно есть поле text или name
            for key in ("text", "name", "value"):
                branch_val = value.get(key)
                if isinstance(branch_val, str) and branch_val.strip():
                    return self._normalize_branch_name(branch_val.strip())
        
        if isinstance(value, str):
            return self._normalize_branch_name(value.strip())
        
        if debug:
            print(f"⚠️  DEBUG: Филиал не извлечен из поля {field_id}")
        
        return "Неизвестный филиал"
    
    def _extract_lesson_score(self, task_fields: List[Dict[str, Any]], field_id: int) -> Optional[float]:
        """
        Извлекает оценку за урок (поле 34 или другие числовые поля).
        
        Поддерживает:
        - Числовые поля (number)
        - Текстовые поля с числом (text)
        - Поля выбора с числовым значением (multiple_choice)
        """
        value = self._get_field_value(task_fields, field_id)
        
        if value is None:
            return None
        
        # Если число
        if isinstance(value, (int, float)):
            return float(value)
        
        # Если строка с числом
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            try:
                return float(value)
            except ValueError:
                return None
        
        # Если объект (для multiple_choice, number и других типов)
        if isinstance(value, dict):
            # Для multiple_choice: проверяем choice_names
            choice_names = value.get("choice_names")
            if isinstance(choice_names, list) and len(choice_names) > 0:
                choice = choice_names[0]
                if isinstance(choice, (int, float)):
                    return float(choice)
                if isinstance(choice, str):
                    try:
                        return float(choice.strip())
                    except ValueError:
                        pass
            
            # Для number/text: стандартные поля
            for key in ("value", "number", "text"):
                val = value.get(key)
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    try:
                        return float(val.strip())
                    except ValueError:
                        continue
        
        return None
    
    def _is_teacher_excluded(self, teacher_name: str) -> bool:
        """Проверяет, исключен ли преподаватель из отчета."""
        excluded_list = self.config.get("excluded_teachers", [])
        
        # Точное совпадение
        if teacher_name in excluded_list:
            return True
        
        # Частичное совпадение (фамилия входит в имя)
        teacher_lower = teacher_name.lower()
        for excluded_name in excluded_list:
            if excluded_name.lower() in teacher_lower:
                return True
        
        return False
    
    def _is_branch_excluded(self, branch_name: str) -> bool:
        """Проверяет, исключен ли филиал из отчета."""
        excluded_list = self.config.get("excluded_branches", [])
        
        # Точное совпадение
        if branch_name in excluded_list:
            return True
        
        # Частичное совпадение
        branch_lower = branch_name.lower()
        for excluded_name in excluded_list:
            if excluded_name.lower() in branch_lower:
                return True
        
        return False
    
    def _is_in_period(self, task: Dict[str, Any]) -> bool:
        """Проверяет, попадает ли задача в указанный период."""
        period_config = self.config.get("period", {})
        
        # Если фильтр по периоду отключен
        if not period_config.get("enabled", False):
            return True
        
        # Получаем дату создания задачи
        create_date_str = task.get("create_date")
        if not create_date_str:
            return True  # Если даты нет - включаем
        
        try:
            # Парсим дату создания (формат ISO: 2024-09-01T10:30:00Z)
            task_date = datetime.fromisoformat(create_date_str.replace('Z', '+00:00'))
            
            # Парсим даты из конфига (делаем timezone-aware)
            start_date_str = period_config.get("start_date")
            end_date_str = period_config.get("end_date")
            
            if start_date_str:
                # Создаем timezone-aware дату
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
                # Делаем naive datetime timezone-aware (используем UTC)
                from datetime import timezone
                start_date = start_date.replace(tzinfo=timezone.utc)
                if task_date < start_date:
                    return False
            
            if end_date_str:
                # Создаем timezone-aware дату
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                # Добавляем 1 день чтобы включить весь конечный день
                end_date = end_date.replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
                if task_date > end_date:
                    return False
            
            return True
        
        except (ValueError, AttributeError) as e:
            print(f"⚠️ Ошибка парсинга даты: {e}")
            return True  # В случае ошибки включаем задачу
    
    def _extract_catalog_value(self, task_fields: List[Dict[str, Any]], field_id: int) -> str:
        """Извлекает значение из поля справочника (категория/программа/курс)."""
        value = self._get_field_value(task_fields, field_id)
        
        if isinstance(value, dict):
            # Проверяем choice_names для справочника выбора
            choice_names = value.get("choice_names")
            if isinstance(choice_names, list) and len(choice_names) > 0:
                return str(choice_names[0])
            
            # Проверяем массив values
            values = value.get("values")
            if isinstance(values, list) and len(values) > 0:
                return str(values[0])
            
            # Проверяем rows
            rows = value.get("rows")
            if isinstance(rows, list) and len(rows) > 0 and isinstance(rows[0], list) and len(rows[0]) > 0:
                return str(rows[0][0])
            
            # Проверяем text, name, value
            for key in ("text", "name", "value"):
                cat_val = value.get(key)
                if isinstance(cat_val, str) and cat_val.strip():
                    return cat_val.strip()
        
        if isinstance(value, str):
            return value.strip()
        
        return "Не указано"
    
    async def analyze_form_697550(self) -> None:
        """Анализ формы 697550 (проверка уроков)."""
        form_id = self.config.get("form_id", 697550)
        include_archived = self.config.get("include_archived", True)
        
        print(f"Анализ формы {form_id} (проверка уроков)...")
        period_config = self.config.get('period', {})
        if period_config.get('enabled'):
            print(f"  • Период фильтрации: Включен ({period_config.get('start_date')} → {period_config.get('end_date')})")
        else:
            print(f"  • Период фильтрации: Отключен")
        print(f"  • Исключенных преподавателей: {len(self.config.get('excluded_teachers', []))}")
        print(f"  • Исключенных филиалов: {len(self.config.get('excluded_branches', []))}")
        print(f"  • Диапазон оценок: {self.config.get('min_score', 1)}-{self.config.get('max_score', 10)}")
        print()
        
        branch_field_id = 47
        teacher_field_id = 48
        score_field_id = 34
        category_field_id = 45
        program_field_id = 37
        course_field_id = 52
        
        task_count = 0
        
        async for task in self.client.iter_register_tasks(form_id, include_archived=include_archived):
            task_count += 1
            if task_count % 100 == 0:
                print(f"Обработано {task_count} задач формы 697550...")
            
            task_fields = task.get("fields", [])
            task_id = task.get("id")
            
            self.total_tasks_processed += 1
            
            # Трекаем даты для статистики
            create_date_str = task.get("create_date")
            if create_date_str:
                try:
                    task_date = datetime.fromisoformat(create_date_str.replace('Z', '+00:00'))
                    if self.earliest_date is None or task_date < self.earliest_date:
                        self.earliest_date = task_date
                    if self.latest_date is None or task_date > self.latest_date:
                        self.latest_date = task_date
                except Exception:
                    pass
            
            # Фильтр 1: Проверка периода
            if not self._is_in_period(task):
                self.excluded_by_period_count += 1
                continue
            
            # Извлекаем данные
            branch_name = self._extract_branch_name(task_fields, branch_field_id)
            
            # Считаем "Неизвестный филиал" (только для статистики)
            if branch_name == "Неизвестный филиал":
                self.unknown_branch_count += 1
            
            teacher_name = self._extract_teacher_name(task_fields, teacher_field_id)
            lesson_score = self._extract_lesson_score(task_fields, score_field_id)
            category = self._extract_catalog_value(task_fields, category_field_id)
            program = self._extract_catalog_value(task_fields, program_field_id)
            course = self._extract_catalog_value(task_fields, course_field_id)
            
            # Фильтр 2: Проверка исключенных преподавателей
            if self._is_teacher_excluded(teacher_name):
                self.excluded_teachers_count += 1
                continue
            
            # Фильтр 3: Проверка исключенных филиалов
            if self._is_branch_excluded(branch_name):
                self.excluded_branches_count += 1
                continue
            
            # Фильтр 4: Валидация оценки (диапазон из конфига)
            min_score = self.config.get("min_score", 1)
            max_score = self.config.get("max_score", 10)
            
            if lesson_score is None or lesson_score < min_score or lesson_score > max_score:
                self.invalid_score_count += 1
                continue  # Пропускаем задачи с невалидной оценкой
            
            self.valid_tasks_count += 1
            
            # Обновляем статистику преподавателя
            if teacher_name not in self.teachers_stats:
                self.teachers_stats[teacher_name] = LessonCheckStats(teacher_name)
            
            teacher_stats = self.teachers_stats[teacher_name]
            teacher_stats.add_check(
                score=lesson_score,
                branch=branch_name,
                category=category,
                program=program,
                course=course,
                task_id=task_id
            )
            
            # Обновляем статистику филиала
            if branch_name not in self.branches_stats:
                self.branches_stats[branch_name] = BranchLessonStats(branch_name)
            
            branch_stats = self.branches_stats[branch_name]
            branch_stats.add_check(lesson_score, teacher_name)
        
        print(f"Завершен анализ формы {form_id}.")
        print(f"  • Обработано задач: {task_count}")
        
        # Выводим период реальных дат
        if self.earliest_date and self.latest_date:
            earliest_str = self.earliest_date.strftime("%Y-%m-%d")
            latest_str = self.latest_date.strftime("%Y-%m-%d")
            print(f"  • Период задач (create_date): {earliest_str} → {latest_str}")
        
        print(f"  • Валидных проверок: {self.valid_tasks_count}")
        print(f"  • Исключено по периоду: {self.excluded_by_period_count}")
        print(f"  • Исключено преподавателей: {self.excluded_teachers_count}")
        print(f"  • Исключено филиалов: {self.excluded_branches_count}")
        print(f"  • Невалидных оценок: {self.invalid_score_count}")
        print(f"  • 'Неизвестный филиал': {self.unknown_branch_count}")
        print()
    
    def create_excel_reports(self, filename: str = "lesson_checks_report.xlsx") -> None:
        """Создает полный Excel файл с 3 вкладками."""
        print(f"Создание Excel отчета: {filename}")
        
        # Создаем один файл с тремя листами
        wb = Workbook()
        
        # Удаляем дефолтный лист
        wb.remove(wb.active)
        
        # Вкладка 1: Рейтинг филиалов по средней оценке
        print("Создание вкладки 'Рейтинг филиалов по средней оценке'...")
        self._create_branches_rating_sheet(wb)
        
        # Вкладка 2: Детали по филиалам
        print("Создание вкладки 'Детали по филиалам'...")
        self._create_branches_details_sheet(wb)
        
        # Вкладка 3: Рейтинги по категориям, программам и курсам
        print("Создание вкладки 'Рейтинг по категориям, программам и курсам'...")
        self._create_rankings_sheet(wb)
        
        # Сохраняем файл
        wb.save(filename)
        print(f"✅ Отчет сохранен: {filename}")
    
    def _create_branches_rating_sheet(self, wb: Workbook) -> None:
        """Создает вкладку 'Рейтинг филиалов по средней оценке'."""
        ws = wb.create_sheet("Рейтинг филиалов")
        
        # Заголовки
        headers = [
            "🏢 Филиал",
            "📊 Количество проверок",
            "⭐ Средняя оценка",
            "🏆 Приз"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Сортируем филиалы по средней оценке (убывание)
        sorted_branches = sorted(
            self.branches_stats.values(),
            key=lambda x: x.average_score,
            reverse=True
        )
        
        # Призы для топ-3 из конфига
        prizes = self.config.get("prizes", {}).get("branch_top3", [
            "💰 15 000 руб.",
            "💰 10 000 руб.",
            "💰 5 000 руб."
        ])
        
        # Данные по филиалам
        row = 2
        for i, branch_stats in enumerate(sorted_branches):
            prize = prizes[i] if i < 3 else ""
            
            ws.cell(row=row, column=1, value=branch_stats.name)
            ws.cell(row=row, column=2, value=branch_stats.total_checks)
            ws.cell(row=row, column=3, value=round(branch_stats.average_score, 2))
            ws.cell(row=row, column=4, value=prize)
            
            # Выделяем топ-3 желтым
            if i < 3:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFD700", end_color="FFD700", fill_type="solid"
                    )
            
            row += 1
        
        # Автоширина колонок
        self._adjust_column_widths(ws)
    
    def _create_branches_details_sheet(self, wb: Workbook) -> None:
        """Создает вкладку 'Детали по филиалам' с отдельными таблицами для каждого филиала."""
        ws = wb.create_sheet("Детали по филиалам")
        
        row = 1
        
        # Сортируем филиалы по средней оценке
        sorted_branches = sorted(
            self.branches_stats.values(),
            key=lambda x: x.average_score,
            reverse=True
        )
        
        for branch_idx, branch_stats in enumerate(sorted_branches):
            branch_name = branch_stats.name
            branch_avg = branch_stats.average_score
            
            # Заголовок филиала (крупный, жирный)
            branch_header = f"🏢 {branch_name}  •  Средний балл: {round(branch_avg, 2)}"
            cell = ws.cell(row=row, column=1, value=branch_header)
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left")
            
            # Объединяем ячейки для заголовка (4 колонки)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
            
            # Заголовки таблицы преподавателей
            headers = [
                "👨‍🏫 Преподаватель",
                "📊 Количество проверок",
                "⭐ Средняя оценка",
                "📈 Ранг"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            row += 1
            
            # Получаем всех преподавателей этого филиала
            teachers_in_branch = []
            
            for teacher_name, teacher_stats in self.teachers_stats.items():
                scores_in_branch = teacher_stats.checks_by_branch.get(branch_name, [])
                if scores_in_branch:
                    avg_score = sum(scores_in_branch) / len(scores_in_branch)
                    teachers_in_branch.append({
                        "name": teacher_name,
                        "checks": len(scores_in_branch),
                        "avg_score": avg_score
                    })
            
            # Сортируем преподавателей по средней оценке (убывание)
            teachers_in_branch.sort(key=lambda x: x["avg_score"], reverse=True)
            
            # Выводим преподавателей
            for rank, teacher_info in enumerate(teachers_in_branch, 1):
                ws.cell(row=row, column=1, value=teacher_info["name"])
                ws.cell(row=row, column=2, value=teacher_info["checks"])
                ws.cell(row=row, column=3, value=round(teacher_info["avg_score"], 2))
                ws.cell(row=row, column=4, value=rank)
                
                # Выделяем топ-3 преподавателей в филиале желтым
                if rank <= 3:
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFD700", end_color="FFD700", fill_type="solid"
                        )
                
                row += 1
            
            # Пустая строка между филиалами
            row += 1
        
        # Автоширина колонок
        self._adjust_column_widths(ws)
    
    def _create_rankings_sheet(self, wb: Workbook) -> None:
        """Создает вкладку 'Рейтинг по категориям, программам и курсам'."""
        ws = wb.create_sheet("Рейтинги")
        
        row = 1
        
        # Раздел 1: Рейтинг по категориям
        row = self._add_ranking_section(
            ws, row, 
            "📊 РЕЙТИНГ ПО КАТЕГОРИЯМ",
            "checks_by_category"
        )
        row += 2
        
        # Раздел 2: Рейтинг по программам
        row = self._add_ranking_section(
            ws, row,
            "📊 РЕЙТИНГ ПО ПРОГРАММАМ",
            "checks_by_program"
        )
        row += 2
        
        # Раздел 3: Рейтинг по курсам
        row = self._add_ranking_section(
            ws, row,
            "📊 РЕЙТИНГ ПО КУРСАМ",
            "checks_by_course"
        )
        
        # Автоширина колонок
        self._adjust_column_widths(ws)
    
    def _add_ranking_section(self, ws, start_row: int, section_title: str, attribute_name: str) -> int:
        """Добавляет секцию рейтинга (по категориям/программам/курсам)."""
        current_row = start_row
        
        # Заголовок секции
        ws.cell(row=current_row, column=1, value=section_title)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="0066CC")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2
        
        # Собираем все группы (категории/программы/курсы)
        groups = defaultdict(list)
        
        for teacher_name, teacher_stats in self.teachers_stats.items():
            checks_by_group = getattr(teacher_stats, attribute_name)
            
            for group_name, scores in checks_by_group.items():
                if scores:
                    avg_score = sum(scores) / len(scores)
                    groups[group_name].append({
                        "teacher": teacher_name,
                        "avg_score": avg_score,
                        "checks": len(scores)
                    })
        
        # Для каждой группы выводим рейтинг
        for group_name in sorted(groups.keys()):
            # Заголовок группы
            ws.cell(row=current_row, column=1, value=f"Группа: {group_name}")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            current_row += 1
            
            # Заголовки таблицы
            headers = ["👨‍🏫 Преподаватель", "⭐ Средняя оценка", "📊 Количество проверок"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            current_row += 1
            
            # Сортируем преподавателей по средней оценке
            teachers_list = sorted(groups[group_name], key=lambda x: x["avg_score"], reverse=True)
            
            # Выводим преподавателей
            for rank, teacher_info in enumerate(teachers_list):
                ws.cell(row=current_row, column=1, value=teacher_info["teacher"])
                ws.cell(row=current_row, column=2, value=round(teacher_info["avg_score"], 2))
                ws.cell(row=current_row, column=3, value=teacher_info["checks"])
                
                # Выделяем топ-3 желтым
                if rank < 3:
                    for col in range(1, 4):
                        ws.cell(row=current_row, column=col).fill = PatternFill(
                            start_color="FFD700", end_color="FFD700", fill_type="solid"
                        )
                
                current_row += 1
            
            # Пустая строка между группами
            current_row += 1
        
        return current_row
    
    def _adjust_column_widths(self, ws) -> None:
        """Настраивает автоширину колонок."""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            if column_letter:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    async def run_analysis(self) -> None:
        """Запускает полный анализ данных."""
        # Получаем настройки output из конфига
        output_config = self.config.get("output", {})
        reports_dir = Path(output_config.get("directory", "reports"))
        reports_dir.mkdir(exist_ok=True)
        
        report_name = self.config.get("report_name", "Отчет по проверкам уроков")
        print(f"{'=' * 80}")
        print(f"  {report_name}")
        print(f"{'=' * 80}")
        print(f"Время начала: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()
        
        # Анализируем форму
        await self.analyze_form_697550()
        
        # Выводим краткую статистику
        print()
        print("=" * 80)
        print("ИТОГОВАЯ СТАТИСТИКА")
        print("=" * 80)
        print(f"Всего филиалов в отчете: {len(self.branches_stats)}")
        print(f"Всего преподавателей в отчете: {len(self.teachers_stats)}")
        print(f"Всего валидных проверок: {self.valid_tasks_count}")
        print()
        print("ИСКЛЮЧЕНО:")
        print(f"  • По периоду: {self.excluded_by_period_count}")
        print(f"  • Преподавателей: {self.excluded_teachers_count}")
        print(f"  • Филиалов: {self.excluded_branches_count}")
        print(f"  • Невалидных оценок: {self.invalid_score_count}")
        print("=" * 80)
        print()
        
        # Создаем Excel отчет
        filename_prefix = output_config.get("filename_prefix", "lesson_checks_report")
        add_timestamp = output_config.get("add_timestamp", True)
        
        if add_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{reports_dir}/{filename_prefix}_{timestamp}.xlsx"
        else:
            filename = f"{reports_dir}/{filename_prefix}.xlsx"
        
        self.create_excel_reports(filename)
        
        print(f"\n✅ Анализ завершен: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"📊 Отчет сохранен: {filename}")


async def main():
    """Главная функция скрипта."""
    try:
        analyzer = LessonChecksAnalyzer()
        await analyzer.run_analysis()
    except KeyboardInterrupt:
        print("\nОтмена выполнения пользователем.")
        sys.exit(1)
    except Exception as e:
        print(f"Ошибка выполнения: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    # Проверяем переменные окружения
    if not os.getenv("PYRUS_LOGIN") or not os.getenv("PYRUS_SECURITY_KEY"):
        print("ОШИБКА: Не установлены переменные окружения PYRUS_LOGIN и PYRUS_SECURITY_KEY")
        print("Убедитесь, что .env файл настроен корректно.")
        sys.exit(1)
    
    asyncio.run(main())

